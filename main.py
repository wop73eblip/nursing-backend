from fastapi import FastAPI, HTTPException, Depends, Request
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import StreamingResponse
from fastapi.security import HTTPBearer, HTTPAuthorizationCredentials
from pydantic import BaseModel
from typing import Optional, List
import bcrypt as _bcrypt
from jose import JWTError, jwt
from datetime import datetime, timedelta, date as date_type
import math, os, io
from urllib.parse import quote
from dotenv import load_dotenv
from supabase import create_client, Client
from ortools.sat.python import cp_model
import openpyxl
from openpyxl.styles import Font, Border, Side, Alignment, PatternFill

load_dotenv()

app = FastAPI(title="護理排班系統 API")

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=False,
    allow_methods=["*"],
    allow_headers=["*"],
)

SECRET_KEY = os.getenv("SECRET_KEY", "fallback-secret-key")
ALGORITHM = "HS256"
ACCESS_TOKEN_EXPIRE_MINUTES = 60 * 8

SUPABASE_URL = os.getenv("SUPABASE_URL")
SUPABASE_SERVICE_KEY = os.getenv("SUPABASE_SERVICE_KEY")
supabase: Client = create_client(SUPABASE_URL, SUPABASE_SERVICE_KEY)

security = HTTPBearer()


# ── 資料模型
class LoginRequest(BaseModel):
    uid: str
    password: str

class Token(BaseModel):
    access_token: str
    token_type: str
    role: str
    name: str
    uid: str

class UserCreate(BaseModel):
    uid: str
    password: str
    name: str
    role: str        # nurse, dual, admin, superadmin
    level: str       # leader, second, member
    attr: str        # 輪班屬性
    halftime: bool = False
    admin_staff: bool = False   # 行政人員：可預班但不參與一鍵生成
    is_trainee: bool = False     # 新人：照排班規則排、但不計臨床人數、跟隨導師
    mentor_uid: Optional[str] = None   # 導師 uid（空＝沒選導師）
    note: str = ""
    sort_order: Optional[int] = None

class UserPatch(BaseModel):
    name: Optional[str] = None
    role: Optional[str] = None
    level: Optional[str] = None
    attr: Optional[str] = None
    halftime: Optional[bool] = None
    admin_staff: Optional[bool] = None
    is_trainee: Optional[bool] = None
    mentor_uid: Optional[str] = None
    note: Optional[str] = None
    sort_order: Optional[int] = None

class AdminResetPassword(BaseModel):
    new_password: str

class ChangePassword(BaseModel):
    old_password: str
    new_password: str

class ShiftUpdate(BaseModel):
    nurse_uid: str
    date: str
    shift: Optional[str] = None

class RulesUpdate(BaseModel):
    rules: dict

class GameSaveUpdate(BaseModel):
    data: dict        # 遊戲進度（場景、旗標、道具…），整包存 jsonb

class GameContentUpdate(BaseModel):
    data: dict        # 遊戲內容（對話、道具文字…），後台編輯，整包存 jsonb

class GameMessageCreate(BaseModel):
    text: str         # 玩家在遊戲留言板送出的內容


# ── 登入防暴力破解：同一帳號+IP 短時間內試錯太多次即暫時鎖定
_LOGIN_FAILS: dict[str, list[float]] = {}   # key -> 失敗時間戳列表
LOGIN_MAX_FAILS = 8            # 視窗內允許的失敗次數
LOGIN_WINDOW_SEC = 600         # 統計視窗（10 分鐘）
LOGIN_LOCK_SEC = 600           # 觸發後鎖定時間（10 分鐘）

def _login_key(uid: str, ip: str) -> str:
    return f"{(uid or '').strip().lower()}|{ip}"

def _login_is_locked(key: str) -> int:
    """回傳剩餘鎖定秒數（0＝未鎖定）。"""
    import time
    now = time.time()
    fails = [t for t in _LOGIN_FAILS.get(key, []) if now - t < LOGIN_WINDOW_SEC]
    _LOGIN_FAILS[key] = fails
    if len(fails) >= LOGIN_MAX_FAILS:
        remain = int(LOGIN_LOCK_SEC - (now - fails[-1]))
        return max(0, remain)
    return 0

def _login_record_fail(key: str) -> None:
    import time
    _LOGIN_FAILS.setdefault(key, []).append(time.time())

def _login_clear(key: str) -> None:
    _LOGIN_FAILS.pop(key, None)


# ── 工具函數
def verify_password(plain: str, hashed: str) -> bool:
    return _bcrypt.checkpw(plain.encode(), hashed.encode())

def get_password_hash(password: str) -> str:
    return _bcrypt.hashpw(password.encode(), _bcrypt.gensalt()).decode()

def create_access_token(data: dict) -> str:
    to_encode = data.copy()
    expire = datetime.utcnow() + timedelta(minutes=ACCESS_TOKEN_EXPIRE_MINUTES)
    to_encode.update({"exp": expire})
    return jwt.encode(to_encode, SECRET_KEY, algorithm=ALGORITHM)

def get_current_user(credentials: HTTPAuthorizationCredentials = Depends(security)):
    try:
        payload = jwt.decode(credentials.credentials, SECRET_KEY, algorithms=[ALGORITHM])
        uid = payload.get("sub")
        if uid is None:
            raise HTTPException(status_code=401, detail="無效的 Token")
        return payload
    except JWTError:
        raise HTTPException(status_code=401, detail="Token 已過期或無效")

def require_roles(*roles):
    def checker(current_user: dict = Depends(get_current_user)):
        if current_user.get("role") not in roles:
            raise HTTPException(status_code=403, detail="權限不足")
        return current_user
    return checker


# ── 路由
@app.get("/")
def root():
    return {"message": "護理排班系統 API 運行中", "version": "2.0.0"}


@app.post("/auth/login", response_model=Token)
def login(request: LoginRequest, http_request: Request):
    uid_in = (request.uid or "").strip()

    # 取得來源 IP（Railway 在反向代理後，優先讀 X-Forwarded-For）
    fwd = http_request.headers.get("x-forwarded-for", "")
    client_ip = (fwd.split(",")[0].strip() if fwd else (http_request.client.host if http_request.client else "unknown"))
    rl_key = _login_key(uid_in, client_ip)

    # 防暴力破解：鎖定中直接拒絕
    remain = _login_is_locked(rl_key)
    if remain > 0:
        raise HTTPException(status_code=429, detail=f"登入嘗試過於頻繁，請於 {remain // 60 + 1} 分鐘後再試")

    # 帳號比對：取全部使用者於程式端做「大小寫不敏感的精確比對」，
    # 避免 ilike 把 % _ 當萬用字元造成的比對漏洞（uid 表小，效能無虞）
    res = supabase.table("users").select("uid, password_hash, role, name").execute()
    user = next((u for u in (res.data or []) if (u.get("uid") or "").lower() == uid_in.lower()), None)

    if not user or not verify_password(request.password, user["password_hash"]):
        _login_record_fail(rl_key)
        raise HTTPException(status_code=401, detail="喔喔!! 帳號或密碼錯了")

    _login_clear(rl_key)   # 成功後清除失敗紀錄
    token = create_access_token({
        "sub": user["uid"],
        "role": user["role"],
        "name": user["name"],
    })
    return {
        "access_token": token,
        "token_type": "bearer",
        "role": user["role"],
        "name": user["name"],
        "uid": user["uid"],
    }


@app.get("/users")
def get_users(current_user: dict = Depends(get_current_user)):
    try:
        res = supabase.table("users").select(
            "uid, name, role, level, attr, halftime, admin_staff, is_trainee, mentor_uid, note, sort_order, created_at"
        ).order("sort_order").order("created_at").execute()
    except Exception:
        res = supabase.table("users").select(
            "uid, name, role, level, attr, halftime, admin_staff, is_trainee, mentor_uid, note, created_at"
        ).order("created_at").execute()
    return {"users": res.data}


@app.post("/users", status_code=201)
def create_user(
    user: UserCreate,
    current_user: dict = Depends(require_roles("admin", "superadmin", "dual")),
):
    if user.role == "superadmin" and current_user.get("role") != "superadmin":
        raise HTTPException(status_code=403, detail="只有超級管理員可新增超級管理員帳號")

    # 帳號統一轉大寫儲存（顯示/班表/匯出一致），登入與唯一性檢查皆大小寫不敏感
    uid_new = (user.uid or "").strip().upper()
    if not uid_new:
        raise HTTPException(status_code=400, detail="帳號 UID 不可為空")
    # 程式端做大小寫不敏感精確比對（避免 ilike 萬用字元問題）
    all_uids = supabase.table("users").select("uid").execute()
    dup = next((u for u in (all_uids.data or []) if (u.get("uid") or "").lower() == uid_new.lower()), None)
    if dup:
        raise HTTPException(status_code=400, detail=f"此帳號 UID 已存在（{dup['uid']}，不分大小寫）")

    # 計算 sort_order：接在最後
    if user.sort_order is None:
        cnt = supabase.table("users").select("uid", count="exact").execute()
        sort_order = (cnt.count or 0) + 1
    else:
        sort_order = user.sort_order

    supabase.table("users").insert({
        "uid": uid_new,
        "password_hash": get_password_hash(user.password),
        "name": user.name,
        "role": user.role,
        "level": user.level,
        "attr": user.attr,
        "halftime": user.halftime,
        "admin_staff": user.admin_staff,
        "is_trainee": user.is_trainee,
        "mentor_uid": user.mentor_uid or None,
        "note": user.note,
        "sort_order": sort_order,
    }).execute()
    return {"message": "帳號建立成功", "uid": uid_new}


@app.patch("/users/{uid}")
def patch_user(
    uid: str,
    body: UserPatch,
    current_user: dict = Depends(get_current_user),
):
    requester = current_user.get("sub")
    requester_role = current_user.get("role")

    if requester != uid and requester_role not in ["admin", "superadmin", "dual"]:
        raise HTTPException(status_code=403, detail="權限不足")

    update_data = body.model_dump(exclude_none=True)

    # 角色變更權限邏輯
    if "role" in update_data:
        new_role = update_data["role"]
        if requester_role == "superadmin":
            pass  # 超管無限制
        elif requester_role in ["admin", "dual"]:
            # 管理員可操作所有非超管帳號，但不能升為超管
            target_res = supabase.table("users").select("role").eq("uid", uid).single().execute()
            target_role = target_res.data.get("role") if target_res.data else None
            if target_role == "superadmin":
                raise HTTPException(status_code=403, detail="無法修改超級管理員帳號")
            if new_role == "superadmin":
                raise HTTPException(status_code=403, detail="只有超級管理員可設定超級管理員角色")
        else:
            raise HTTPException(status_code=403, detail="權限不足")

    if update_data:
        supabase.table("users").update(update_data).eq("uid", uid).execute()
    return {"message": "更新成功"}


@app.post("/auth/change-password")
def change_password(
    body: ChangePassword,
    current_user: dict = Depends(get_current_user),
):
    uid = current_user.get("sub")
    res = supabase.table("users").select("password_hash").eq("uid", uid).single().execute()
    if not res.data or not verify_password(body.old_password, res.data["password_hash"]):
        raise HTTPException(status_code=400, detail="目前密碼不正確")
    if len(body.new_password) < 4:
        raise HTTPException(status_code=400, detail="新密碼至少 4 個字元")
    supabase.table("users").update({
        "password_hash": get_password_hash(body.new_password)
    }).eq("uid", uid).execute()
    return {"message": "密碼已變更"}


@app.post("/users/{uid}/reset-password")
def reset_password(
    uid: str,
    body: AdminResetPassword,
    current_user: dict = Depends(require_roles("admin", "superadmin")),
):
    if len(body.new_password) < 4:
        raise HTTPException(status_code=400, detail="密碼至少 4 個字元")
    supabase.table("users").update({
        "password_hash": get_password_hash(body.new_password)
    }).eq("uid", uid).execute()
    return {"message": "密碼已重設"}



@app.delete("/users/{uid}")
def delete_user(
    uid: str,
    current_user: dict = Depends(require_roles("admin", "superadmin")),
):
    if uid == current_user.get("sub"):
        raise HTTPException(status_code=400, detail="無法刪除自己的帳號")
    supabase.table("users").delete().eq("uid", uid).execute()
    return {"message": "帳號已刪除"}


@app.post("/users/reorder")
def reorder_users(
    order: List[str],   # list of uid in new order
    current_user: dict = Depends(require_roles("admin", "superadmin", "dual")),
):
    for i, uid in enumerate(order):
        supabase.table("users").update({"sort_order": i}).eq("uid", uid).execute()
    return {"message": "排序已更新"}


@app.get("/schedule")
def get_schedule(
    year: int,
    month: int,
    current_user: dict = Depends(get_current_user),
):
    start = f"{year}-{month:02d}-01"
    end = f"{year+1}-01-01" if month == 12 else f"{year}-{month+1:02d}-01"
    res = supabase.table("shifts").select("*").gte("date", start).lt("date", end).execute()
    return {"schedule": res.data}


@app.post("/schedule/shift")
def update_shift(
    update: ShiftUpdate,
    current_user: dict = Depends(get_current_user),
):
    role = current_user.get("role")
    uid = current_user.get("sub")

    if role == "nurse" and update.nurse_uid != uid:
        raise HTTPException(status_code=403, detail="只能修改自己的預班")

    existing = supabase.table("shifts").select("id").eq("nurse_uid", update.nurse_uid).eq("date", update.date).execute()

    if existing.data:
        supabase.table("shifts").update({
            "shift": update.shift,
            "confirmed": False,
            "updated_by": uid,
            "updated_at": datetime.utcnow().isoformat(),
        }).eq("nurse_uid", update.nurse_uid).eq("date", update.date).execute()
    else:
        supabase.table("shifts").insert({
            "code": f"{update.nurse_uid}_{update.date}",   # unique per nurse+date
            "label": update.shift or "",
            "nurse_uid": update.nurse_uid,
            "date": update.date,
            "shift": update.shift,
            "confirmed": False,
            "updated_by": uid,
        }).execute()

    try:
        supabase.table("shift_logs").insert({
            "nurse_uid": update.nurse_uid,
            "date": update.date,
            "shift": update.shift,
            "changed_by": uid,            # 舊欄位 NOT NULL 相容
            "operator_uid": uid,
            "operator_role": role,
            "action": "edit",
        }).execute()
    except Exception:
        pass  # log 失敗不影響主流程

    return {"message": "班別更新成功"}


class ShiftBatchItem(BaseModel):
    nurse_uid: str
    date: str
    shift: Optional[str] = None


@app.post("/schedule/shifts/batch")
def batch_update_shifts(
    updates: List[ShiftBatchItem],
    current_user: dict = Depends(get_current_user),
):
    """批次寫入多格班別（一次 API 呼叫取代多次單格呼叫）"""
    if not updates:
        return {"message": "無資料", "updated": 0}

    role = current_user.get("role")
    uid  = current_user.get("sub")

    # 護理師只能寫自己的格子
    if role == "nurse":
        if any(u.nurse_uid != uid for u in updates):
            raise HTTPException(status_code=403, detail="只能修改自己的預班")

    # 一次查出所有涉及的 nurse+date 是否已存在
    pairs = [(u.nurse_uid, u.date) for u in updates]
    # 依 nurse_uid 分組查詢，避免 OR 組合過長
    existing_keys: set[str] = set()
    nurse_uids = list({p[0] for p in pairs})
    min_date = min(p[1] for p in pairs)
    max_date = max(p[1] for p in pairs)
    res = supabase.table("shifts").select("nurse_uid, date") \
        .in_("nurse_uid", nurse_uids) \
        .gte("date", min_date).lte("date", max_date).execute()
    for r in (res.data or []):
        existing_keys.add(f"{r['nurse_uid']}_{r['date']}")

    to_insert, to_update_clear, to_update_set = [], [], []
    for u in updates:
        key = f"{u.nurse_uid}_{u.date}"
        if key in existing_keys:
            if u.shift:
                to_update_set.append(u)
            else:
                to_update_clear.append(u)
        else:
            if u.shift:   # shift=None 且不存在 → 無需動作
                to_insert.append(u)

    now = datetime.utcnow().isoformat()

    # 批次 INSERT
    if to_insert:
        supabase.table("shifts").insert([{
            "code": f"{u.nurse_uid}_{u.date}",
            "label": u.shift or "",
            "nurse_uid": u.nurse_uid,
            "date": u.date,
            "shift": u.shift,
            "confirmed": False,
            "updated_by": uid,
        } for u in to_insert]).execute()

    # UPDATE（set shift）
    for u in to_update_set:
        supabase.table("shifts").update({
            "shift": u.shift, "confirmed": False,
            "updated_by": uid, "updated_at": now,
        }).eq("nurse_uid", u.nurse_uid).eq("date", u.date).execute()

    # UPDATE（clear shift → shift=null）
    for u in to_update_clear:
        supabase.table("shifts").update({
            "shift": None, "confirmed": False,
            "updated_by": uid, "updated_at": now,
        }).eq("nurse_uid", u.nurse_uid).eq("date", u.date).execute()

    # 批次寫 log
    try:
        supabase.table("shift_logs").insert([{
            "nurse_uid": u.nurse_uid, "date": u.date, "shift": u.shift,
            "changed_by": uid, "operator_uid": uid,
            "operator_role": role, "action": "edit",
        } for u in updates]).execute()
    except Exception:
        pass

    total = len(to_insert) + len(to_update_set) + len(to_update_clear)
    return {"message": f"批次更新完成", "updated": total}


@app.post("/schedule/confirm")
def confirm_shifts(
    shifts: List[ShiftUpdate],
    current_user: dict = Depends(get_current_user),
):
    uid = current_user.get("sub")
    role = current_user.get("role")

    for s in shifts:
        # 護理師只能確認自己的班
        if role == "nurse" and s.nurse_uid != uid:
            raise HTTPException(status_code=403, detail="只能確認自己的班別")

        existing = supabase.table("shifts").select("id").eq("nurse_uid", s.nurse_uid).eq("date", s.date).execute()
        if existing.data:
            supabase.table("shifts").update({
                "shift": s.shift,
                "confirmed": True,
                "updated_by": uid,
                "updated_at": datetime.utcnow().isoformat(),
            }).eq("nurse_uid", s.nurse_uid).eq("date", s.date).execute()
        else:
            supabase.table("shifts").insert({
                "code": f"{s.nurse_uid}_{s.date}",
                "label": s.shift or "",
                "nurse_uid": s.nurse_uid,
                "date": s.date,
                "shift": s.shift,
                "confirmed": True,
                "updated_by": uid,
            }).execute()

        try:
            supabase.table("shift_logs").insert({
                "nurse_uid": s.nurse_uid,
                "date": s.date,
                "shift": s.shift,
                "changed_by": uid,
                "operator_uid": uid,
                "operator_role": role,
                "action": "confirm",
            }).execute()
        except Exception:
            pass

    return {"message": f"已確認 {len(shifts)} 筆班別"}


@app.post("/schedule/unconfirm")
def unconfirm_shifts(
    shifts: List[ShiftUpdate],
    current_user: dict = Depends(require_roles("admin", "superadmin", "dual")),
):
    uid = current_user.get("sub")
    role = current_user.get("role")
    for s in shifts:
        supabase.table("shifts").update({
            "confirmed": False,
            "updated_by": uid,
            "updated_at": datetime.utcnow().isoformat(),
        }).eq("nurse_uid", s.nurse_uid).eq("date", s.date).execute()
        try:
            supabase.table("shift_logs").insert({
                "nurse_uid": s.nurse_uid,
                "date": s.date,
                "shift": s.shift,
                "changed_by": uid,
                "operator_uid": uid,
                "operator_role": role,
                "action": "unconfirm",
            }).execute()
        except Exception:
            pass
    return {"message": f"已取消確認 {len(shifts)} 筆"}


@app.get("/rules")
def get_rules(current_user: dict = Depends(get_current_user)):
    res = supabase.table("rules").select("*").limit(1).execute()
    if res.data:
        return {"rules": res.data[0].get("data") or {}}
    return {"rules": {}}


@app.get("/login-config")
def get_login_config():
    """公開端點（登入頁未帶 token）：回傳登入畫面自訂內容，未設定則回空值由前端用預設。"""
    res = supabase.table("rules").select("data").limit(1).execute()
    login = {}
    if res.data:
        login = (res.data[0].get("data") or {}).get("login") or {}
    return {
        "title": login.get("title") or "",
        "subtitle": login.get("subtitle") or "",
        "image": login.get("image") or "",   # base64 data URI，空字串代表用預設心電圖圖示
    }


# 登入後首頁的模組卡片預設值（大標/小標/圖片可於後台自訂；enabled 由程式控制、非使用者可改）
DEFAULT_MODULES = [
    {"key": "schedule", "title": "排班系統", "tagline": "不來預班就沒得預班囉～", "enabled": True},
    {"key": "data",     "title": "學習系統", "tagline": "護理訓練小遊戲",           "enabled": True},
]


@app.get("/home-config")
def get_home_config(current_user: dict = Depends(get_current_user)):
    """登入後首頁的模組卡片設定：合併後台自訂（大標/小標/圖片）與預設值。"""
    res = supabase.table("rules").select("data").limit(1).execute()
    saved: dict = {}
    if res.data:
        for m in ((res.data[0].get("data") or {}).get("modules") or []):
            if m.get("key"):
                saved[m["key"]] = m
    modules = []
    for d in DEFAULT_MODULES:
        s = saved.get(d["key"], {})
        modules.append({
            "key": d["key"],
            "title": s.get("title") or d["title"],
            "tagline": s.get("tagline") or d["tagline"],
            "image": s.get("image") or "",   # base64 data URI，空＝用預設圖示
            "enabled": d["enabled"],
        })
    return {"modules": modules}


@app.post("/rules")
def save_rules(
    body: RulesUpdate,
    current_user: dict = Depends(require_roles("admin", "superadmin", "dual")),
):
    existing = supabase.table("rules").select("id", "data").limit(1).execute()
    if existing.data:
        current_data = existing.data[0].get("data") or {}
        incoming = dict(body.rules)
        # modules 特殊處理：依 key 合併（更新有送的、保留沒送的），
        # 讓「排班後台」「學習系統後台」各自只送自己那張卡也不會蓋掉對方。
        if "modules" in incoming:
            cur = {m["key"]: m for m in (current_data.get("modules") or []) if m.get("key")}
            for m in incoming["modules"]:
                if m.get("key"):
                    cur[m["key"]] = m
            incoming["modules"] = list(cur.values())
        merged = {**current_data, **incoming}
        supabase.table("rules").update({
            "data": merged,
            "updated_at": datetime.utcnow().isoformat(),
        }).eq("id", existing.data[0]["id"]).execute()
    else:
        supabase.table("rules").insert({
            "key": "config",   # 舊欄位 NOT NULL 相容
            "value": "{}",
            "data": body.rules,
        }).execute()
    return {"message": "規則已儲存"}


# ── 遊戲存檔：與現有帳號綁定（uid 一律由登入 token 取得，不信任前端傳來的身分）
@app.get("/game/save")
def get_game_save(current_user: dict = Depends(get_current_user)):
    uid = current_user["sub"]
    res = supabase.table("game_saves").select("data").eq("uid", uid).limit(1).execute()
    if res.data:
        return {"data": res.data[0].get("data")}
    return {"data": None}   # 沒存檔＝新玩家


@app.put("/game/save")
def put_game_save(
    body: GameSaveUpdate,
    current_user: dict = Depends(get_current_user),
):
    uid = current_user["sub"]
    supabase.table("game_saves").upsert({
        "uid": uid,
        "data": body.data,
        "updated_at": datetime.utcnow().isoformat(),
    }).execute()
    return {"message": "已儲存"}


@app.delete("/game/save")
def delete_game_save(current_user: dict = Depends(get_current_user)):
    uid = current_user["sub"]
    supabase.table("game_saves").delete().eq("uid", uid).execute()
    return {"message": "已清除"}


# ── 遊戲內容（對話、道具文字）：後台編輯，全遊戲共用一份
@app.get("/game/content")
def get_game_content():
    """公開：遊戲載入時抓最新內容。沒設定過就回空物件，遊戲會用內建預設。"""
    res = supabase.table("game_content").select("data").eq("id", 1).limit(1).execute()
    if res.data:
        return {"data": res.data[0].get("data") or {}}
    return {"data": {}}


@app.post("/game/content")
def save_game_content(
    body: GameContentUpdate,
    current_user: dict = Depends(require_roles("superadmin")),   # 遊戲後台：僅超級管理員
):
    supabase.table("game_content").upsert({
        "id": 1,   # 只存一列，永遠覆蓋
        "data": body.data,
        "updated_at": datetime.utcnow().isoformat(),
    }).execute()
    return {"message": "已儲存"}


# ── 遊戲留言板：玩家送出留言（綁登入帳號）
@app.post("/game/messages", status_code=201)
def post_game_message(
    body: GameMessageCreate,
    current_user: dict = Depends(get_current_user),
):
    text = (body.text or "").strip()
    if not text:
        raise HTTPException(status_code=400, detail="留言不能空白")
    supabase.table("game_messages").insert({
        "uid": current_user["sub"],
        "name": current_user.get("name"),
        "text": text[:500],   # 上限 500 字
    }).execute()
    return {"message": "已送出"}


@app.get("/game/messages")
def get_game_messages(current_user: dict = Depends(get_current_user)):
    # 超級管理員（遊戲後台）看全部；其他人只看自己的
    q = supabase.table("game_messages").select("id, uid, name, text, created_at")
    if current_user.get("role") != "superadmin":
        q = q.eq("uid", current_user["sub"])
    res = q.order("created_at", desc=True).limit(300).execute()
    return {"messages": res.data or []}


@app.delete("/game/messages/{msg_id}")
def delete_game_message(
    msg_id: int,
    current_user: dict = Depends(require_roles("superadmin")),   # 遊戲後台：僅超級管理員
):
    supabase.table("game_messages").delete().eq("id", msg_id).execute()
    return {"message": "已刪除"}


@app.post("/schedule/generate")
def generate_schedule(
    overwrite_confirmed: bool = False,
    profile: str = "balanced",   # smooth=順班優先 / fair=公平優先 / balanced=預設
    current_user: dict = Depends(require_roles("admin", "superadmin", "dual")),
):
    """
    CP-SAT 排班演算法，遵守以下規則：
    硬規則：leader 配置、反向班、每週至少休1天、每週至多兩種班別、連班上限
    軟規則：順班（同種班連排）、固定班不切換、符合各班人數需求
    休假規則：週期首週末不同時休、每週休假上限、一例一休、特休最高順位
    """
    operator_uid = current_user.get("sub")

    # 內建原則：已填班別（含未確認）一律保留，只填寫空白格子
    overwrite_confirmed = False

    # 權重版本：smooth=順班優先（切換懲罰×2）；fair=公平優先（比例/公平懲罰×2）
    SWITCH_MULT = 2 if profile == "smooth" else 1
    FAIR_MULT   = 2 if profile == "fair" else 1

    # 除錯用：DEBUG_SKIP_RULES=rule4,one17,... 可停用特定硬規則群組（生產環境不設定即無作用）
    DEBUG_SKIP = set(filter(None, os.getenv("DEBUG_SKIP_RULES", "").split(",")))

    # ── 讀取規則
    rules_res = supabase.table("rules").select("*").limit(1).execute()
    if not rules_res.data:
        raise HTTPException(400, "請先設定排班規則")
    rules = rules_res.data[0].get("data") or {}

    cycle      = rules.get("cycle", {})
    scheduling = rules.get("scheduling", {})
    ratio      = rules.get("ratio", {})
    ratio_overrides_list = rules.get("ratio_overrides", [])
    ratio_overrides = {o["nurse_uid"]: o["ratio"] for o in ratio_overrides_list}

    start_str = cycle.get("start_date")
    end_str   = cycle.get("end_date")
    if not start_str or not end_str:
        raise HTTPException(400, "請先在「排班週期」設定開始與結束日期")

    start_d = date_type.fromisoformat(start_str)
    end_d   = date_type.fromisoformat(end_str)
    cycle_dates = [
        (start_d + timedelta(days=i)).isoformat()
        for i in range((end_d - start_d).days + 1)
    ]
    n = len(cycle_dates)
    weekdays = [(start_d + timedelta(days=i)).weekday() for i in range(n)]
    # weekday(): Mon=0 … Sat=5, Sun=6

    # ── 規則參數
    max_consec   = int(scheduling.get("max_consecutive_work", 5))
    daily_d      = int(scheduling.get("daily_d", 3))
    daily_e      = int(scheduling.get("daily_e", 3))
    daily_n      = int(scheduling.get("daily_n", 3))
    # 特殊日期人數覆蓋：{date_str: {D:int, E:int, N:int}}
    _special_raw = scheduling.get("special_dates", [])
    special_dates_map: dict[str, dict[str, int]] = {
        sd["date"]: {"D": int(sd.get("d", daily_d)), "E": int(sd.get("e", daily_e)), "N": int(sd.get("n", daily_n))}
        for sd in (_special_raw or []) if sd.get("date")
    }
    # 每日實際需求人數（考慮特殊日期覆蓋）
    day_d = [special_dates_map.get(cycle_dates[t], {}).get("D", daily_d) for t in range(n)]
    day_e = [special_dates_map.get(cycle_dates[t], {}).get("E", daily_e) for t in range(n)]
    day_n = [special_dates_map.get(cycle_dates[t], {}).get("N", daily_n) for t in range(n)]
    total_work_demand = sum(day_d[t] + day_e[t] + day_n[t] for t in range(n))
    no_reverse   = True   # 反向班禁止：固定硬規則，不可關閉（忽略舊設定值）
    restrict_first_weekend = bool(scheduling.get("restrict_first_weekend", True))
    one_in_seven = bool(scheduling.get("one_in_seven", True))   # 一例一休
    lock_first_day       = bool(scheduling.get("lock_first_day", True))
    lock_designated_off  = bool(scheduling.get("lock_designated_off", True))
    weekly_max_off_auto  = int(scheduling.get("weekly_max_off_auto", 2))   # 自動休連續天數上限
    weekly_max_off_total = int(scheduling.get("weekly_max_off_total", 3))  # 每週應休總上限
    holiday_days = int(cycle.get("holiday_days", 0))
    full_off  = min(8 + holiday_days, 13)
    # 半職可上天數（小數點無條件捨去）→ 應休 = 28 - 可上天數
    part_work = math.floor((160 - holiday_days * 8) / 2 / 8)
    part_off  = 28 - part_work

    # 從規則讀取班別定義（前端班別設定儲存的三分類）
    shift_defs = rules.get("shifts", {})
    _rest_defs  = shift_defs.get("rest", [])   # 應休班別（OFF、半）
    _leave_defs = shift_defs.get("off",  [])   # 放假/調整類（V、喪、員⋯）
    _work_defs  = shift_defs.get("work", [])   # 上班類（D、E、N、會、公、書記⋯）

    # 應休代碼集：計入應休天數名額（預設 OFF + 半）
    REST_CODES = {s["code"] for s in _rest_defs if s.get("code")} or {"OFF", "半"}
    # 放假/調整代碼集：最高順位鎖定，不佔應休名額（預設 V、員⋯）
    LEAVE_ADJUST = {s["code"] for s in _leave_defs if s.get("code")} or {"V", "員", "喪", "延休", "補休", "調移"}
    # 行政類上班班別：視同 D（套用 D 的所有規則：反向班、連續上班等），但不佔臨床人力名額
    # （S1/S11/S12 排除）。凡標 admin_only、或非 D/E/N 的上班類（如 會/公/書）皆納入。
    ADMIN_SHIFTS = {
        s["code"] for s in _work_defs
        if s.get("code") and (s.get("admin_only") or s["code"] not in ("D", "E", "N"))
    } or {"會", "公", "書記"}
    # 固定班型索引（用於軟懲罰）
    FIXED_SHIFT_MAP = {"固定D": 0, "固定E": 1, "固定N": 2}

    # ── 讀取護理師（排除行政人員：可預班但不參與生成）
    nurses_res = supabase.table("users").select(
        "uid, name, attr, halftime, level, admin_staff, is_trainee, mentor_uid"
    ).in_("role", ["nurse", "dual"]).order("sort_order").execute()
    nurses = [n for n in (nurses_res.data or []) if not n.get("admin_staff")]
    if not nurses:
        raise HTTPException(400, "尚無護理師帳號")
    M = len(nurses)
    nid = {n["uid"]: i for i, n in enumerate(nurses)}
    # 新人：照排班規則排、但不計臨床人數（S1/S11/S12 排除）；有導師則軟性跟隨其班
    trainee_set = {i for i, nu in enumerate(nurses) if nu.get("is_trainee")}
    mentor_of = {}   # 新人 index → 導師 index（導師須為名單內、非新人）
    for i, nu in enumerate(nurses):
        if nu.get("is_trainee") and nu.get("mentor_uid"):
            mj = nid.get(nu["mentor_uid"])
            if mj is not None and mj not in trainee_set and mj != i:
                mentor_of[i] = mj
    MENTOR_FOLLOW_PENALTY = int(os.getenv("MENTOR_FOLLOW_PENALTY", "3000"))  # 新人每天與導師不同班的軟懲罰（要壓過換班/比例等，才會真的跟著老師）

    # ── 讀取已確認班 / 指定班（含特休、指定休）
    existing_res = supabase.table("shifts").select("nurse_uid, date, shift, confirmed") \
        .gte("date", start_str).lte("date", end_str).execute()
    existing: dict[tuple, dict] = {
        (r["nurse_uid"], r["date"]): r
        for r in (existing_res.data or [])
    }

    # ── helper：計算各護理師每週的 ISO 週邊界
    def week_ranges(dates: list[str]) -> list[tuple[int,int]]:
        """回傳 [(start_i, end_i)] 代表週一到週日的日期索引範圍（週期內）"""
        weeks = []
        i = 0
        while i < len(dates):
            wd = weekdays[i]
            week_start = i - wd  # Mon of that week (may be < 0)
            week_end   = week_start + 6
            # 夾到週期內
            ws = max(0, week_start)
            we = min(len(dates) - 1, week_end)
            weeks.append((ws, we))
            # 跳到下個週一
            i = week_end + 1
        return weeks
    weeks = week_ranges(cycle_dates)

    # ── 計算各護理師應排班次數（依比例）
    def shift_counts(attr: str, work_days: int, nurse_uid: str) -> tuple[int,int,int]:
        if attr == "固定D": return (work_days, 0, 0)
        if attr == "固定E": return (0, work_days, 0)
        if attr == "固定N": return (0, 0, work_days)
        ov = ratio_overrides.get(nurse_uid, {})
        def r(key, default=1): return max(1, int(ov.get(key, ratio.get(key, default))))
        if attr == "輪班DE":
            rd, re = r("D",1), r("E",1); tot = rd+re
            d = round(work_days*rd/tot); return (d, work_days-d, 0)
        if attr == "輪班EN":
            re, rn = r("E",1), r("N",1); tot = re+rn
            e = round(work_days*re/tot); return (0, e, work_days-e)
        if attr == "輪班DN":
            rd, rn = r("D",1), r("N",1); tot = rd+rn
            d = round(work_days*rd/tot); return (d, 0, work_days-d)
        rd, re, rn = r("D",1), r("E",1), r("N",1); tot = rd+re+rn
        d = round(work_days*rd/tot); e = round(work_days*re/tot)
        return (d, e, work_days-d-e)

    # ── 允許班種（依輪班屬性）
    SHIFT_ALLOWED: dict[str, list[str]] = {
        "固定D":   ["D"],
        "固定E":   ["E"],
        "固定N":   ["N"],
        "輪班DE":  ["D","E"],
        "輪班EN":  ["E","N"],
        "輪班DN":  ["D","N"],
        "輪班DEN": ["D","E","N"],
    }
    WORK_SHIFTS = ["D","E","N"]
    SI = {s: i for i, s in enumerate(["D","E","N","OFF"])}  # shift index

    # ── 統一時間軸：history（固定值，不建決策變數）+ decision（待排班，t=0..n-1）
    # HISTORY_DAYS 依規則動態決定，涵蓋所有需要跨界回看的限制（連續上班、反向班 2 日間隔等），
    # 不硬編碼固定天數，之後更換排班區間或規則參數會自動調整。
    HISTORY_DAYS = max(max_consec, weekly_max_off_total, weekly_max_off_auto, 2)
    hist_dates_list = [(start_d - timedelta(days=HISTORY_DAYS - i)).isoformat() for i in range(HISTORY_DAYS)]
    hist_res = supabase.table("shifts").select("nurse_uid, date, shift") \
        .gte("date", hist_dates_list[0]).lte("date", hist_dates_list[-1]).execute()
    hist_raw_by_nurse: dict[str, list[str]] = {}
    for r in (hist_res.data or []):
        uid_r = r["nurse_uid"]
        if uid_r not in hist_raw_by_nurse:
            hist_raw_by_nurse[uid_r] = ["OFF"] * HISTORY_DAYS
        try:
            idx = hist_dates_list.index(r["date"])
            hist_raw_by_nurse[uid_r][idx] = r.get("shift") or "OFF"
        except ValueError:
            pass

    def _hist_shift_to_si(raw: str) -> int:
        """歷史班別代碼 → shift index（0=D,1=E,2=N,3=OFF-equivalent）。
        行政班視同 D；應休/放假調整/OFF 一律視為 OFF-equivalent（不算連續上班）。"""
        s = raw or "OFF"
        if s in ADMIN_SHIFTS:
            return 0
        if s in REST_CODES or s in LEAVE_ADJUST or s == "OFF":
            return 3
        return SI.get(s, 3)

    # hist_si[m] = 固定的歷史 shift index 序列，長度 HISTORY_DAYS（最舊在前，最新在後即 day -1）
    hist_si: dict[int, list[int]] = {}

    # ── CP-SAT 模型
    model = cp_model.CpModel()
    leave_adjust_per_m: dict[int, set[int]] = {}  # 各護理師的 LEAVE_ADJUST 天索引
    locked_si_counts_per_m: dict[int, list[int]] = {}  # 各護理師已鎖定工作班數量 [D,E,N]（供比例硬上限讓路）
    off_slack_vars: list[tuple[int, any]] = []     # (m, slack_var) for shortage warning

    # x[m][t] ∈ {0,1,2,3} → D/E/N/OFF
    x: list[list] = [
        [model.new_int_var(0, 3, f"x_{m}_{t}") for t in range(n)]
        for m in range(M)
    ]
    # bool 變數：is_shift[m][t][s]
    b: list[list[list]] = [
        [[model.new_bool_var(f"b_{m}_{t}_{s}") for s in range(4)] for t in range(n)]
        for m in range(M)
    ]

    def bx(m: int, t: int, s: int):
        """統一時間軸存取：t<0 → 歷史固定值（0/1 常數）；t>=0 → 決策變數 b[m][t][s]。"""
        if t < 0:
            return 1 if hist_si[m][t + HISTORY_DAYS] == s else 0
        return b[m][t][s]

    penalties: list = []
    prefill_conflicts: list[str] = []  # 預填班別與輪班屬性不符的清單
    # 可放寬硬規則的 assumption 開關（正常求解全設為真；無解時用來找 infeasible core）
    assume_reg: list[tuple[object, str]] = []  # (BoolVar, 人類可讀標籤)
    # 行政班格 (m, t)：這些格視同 D 套用 D 規則，但不佔臨床人力（S1/S11/S12 排除）
    admin_cells: set[tuple[int, int]] = set()
    # 全職每週標準休超過 2 天的懲罰（一例一休真正意思：正常週剛好 2 天、不多給）
    WEEKLY_OFF_OVER_PENALTY = int(os.getenv("WEEKLY_OFF_OVER_PENALTY", "500"))

    for m in range(M):
        nurse = nurses[m]
        attr  = nurse.get("attr") or "輪班DEN"
        is_ht = nurse.get("halftime", False)
        lvl   = nurse.get("level", "member")
        uid   = nurse["uid"]

        hist_si[m] = [
            _hist_shift_to_si(s) for s in hist_raw_by_nurse.get(uid, ["OFF"] * HISTORY_DAYS)
        ]

        # x ↔ b 對應
        for t in range(n):
            model.add(x[m][t] == sum(s * b[m][t][s] for s in range(4)))
            model.add_exactly_one(b[m][t])

        # ── 鎖定預填班 / 放假調整類
        locked_off_days_m: set[int] = set()    # 指定 OFF（計入應休名額）
        leave_adjust_days_m: set[int] = set()  # 放假/調整類（不佔應休名額）
        exception_days_m: dict[int, int] = {}   # 屬性衝突的預填格 t→si（跳過允許班種限制）
        nurse_name = nurse.get("name") or uid

        for t, d_str in enumerate(cycle_dates):
            key = (uid, d_str)
            if "locks" in DEBUG_SKIP or key not in existing:
                continue
            row = existing[key]
            # shift 為空（被清除的殘留列）→ 視為空白格，不鎖定
            if not row.get("shift"):
                continue
            orig_shift = row["shift"]
            confirmed = row.get("confirmed", False)

            # 已確認班 + overwrite_confirmed → 跳過（讓 CP-SAT 重排）
            if confirmed and overwrite_confirmed:
                continue

            shift = orig_shift

            # 行政班（書/會/公）：視同 D，套用 D 的所有規則（反向班、連續上班、每週兩種班等），
            # 且算「上班」（計入連續上班、不算應休）。但不佔臨床人力名額 → 記入 admin_cells，
            # 由下方 S1/S11/S12 排除。求解後由還原邏輯改回原行政碼顯示。
            # 不 continue：往下走工作班鎖定（x=D）與固定/允許班種的例外處理。
            if shift in ADMIN_SHIFTS:
                shift = "D"
                admin_cells.add((m, t))

            # 放假/調整類：鎖定為 OFF，不佔應休名額（最高優先）
            if shift in LEAVE_ADJUST:
                model.add(x[m][t] == 3)
                leave_adjust_days_m.add(t)
                continue

            # 應休類（半等）：鎖定為 OFF，計入應休名額
            if shift in REST_CODES and shift != "OFF":
                model.add(x[m][t] == 3)
                locked_off_days_m.add(t)
                continue

            # OFF：鎖定
            if shift == "OFF":
                model.add(x[m][t] == 3)
                locked_off_days_m.add(t)
                continue

            # 工作班（D/E/N 等）：所有預填一律鎖定
            si = SI.get(shift)
            if si is not None:
                model.add(x[m][t] == si)
                locked_si_counts_per_m.setdefault(m, [0, 0, 0])[si] += 1
                # 檢查是否與輪班屬性衝突（輪班類與固定班都追蹤例外日）
                fixed_si_check = FIXED_SHIFT_MAP.get(attr)
                if fixed_si_check is not None:
                    if si != fixed_si_check:
                        exception_days_m[t] = si
                        prefill_conflicts.append(
                            f"{nurse_name}（{attr}）{d_str} 預填 {orig_shift}，不符固定班別"
                        )
                else:
                    allowed_check = SHIFT_ALLOWED.get(attr, WORK_SHIFTS)
                    if shift not in allowed_check:
                        exception_days_m[t] = si
                        prefill_conflicts.append(
                            f"{nurse_name}（{attr}）{d_str} 預填 {orig_shift}，不在允許班別 {'/'.join(allowed_check)}"
                        )

        # ── 允許班種限制（輪班類硬限制；固定班改用軟懲罰；例外日跳過）
        # 自由格仍嚴格維持原班屬允許班種（不含例外班種）。例外日鎖定造成的第三種班種
        # 疊加「硬規則4：每週至多兩種班別」（絕對值，見下方）後，會自然逼迫求解器
        # 讓該週自由格只保留原兩種班別中的其中一種（例如輪班DE遇例外N → 自由格收斂為全D或全E）
        fixed_si = FIXED_SHIFT_MAP.get(attr)
        if fixed_si is None:
            allowed = SHIFT_ALLOWED.get(attr, WORK_SHIFTS)
            allowed_si = set(SI[s] for s in allowed) | {3}
            if "allowed" not in DEBUG_SKIP:
                for t in range(n):
                    if t in exception_days_m:
                        continue  # 屬性衝突的預填日：保留預填值，跳過班種限制
                    for s in range(4):
                        if s not in allowed_si:
                            model.add(b[m][t][s] == 0)

        # ── 休假天數（LEAVE_ADJUST 不計入應休名額）
        la_count = len(leave_adjust_days_m)
        guaranteed_off = part_off if is_ht else full_off
        off_days = guaranteed_off  # 全職用 full_off，半職用 part_off
        off_days = max(0, min(off_days, n - la_count - 1))

        # 應休 OFF（排除 LEAVE_ADJUST）
        free_off = [b[m][t][3] for t in range(n) if t not in leave_adjust_days_m]
        # 軟規則 E：人力不足時允許縮減應休天數（off_slack，最多 -2 天）
        off_slack = model.new_int_var(0, 2, f"off_slack_{m}")
        off_slack_vars.append((m, off_slack))
        if "offdays" not in DEBUG_SKIP:
            model.add(sum(free_off) >= off_days - off_slack)  # 軟下限：最多縮減 2 天
        # 超出應休天數的部分用高懲罰軟約束取代硬上限，避免供需差造成 INFEASIBLE
        over_off = model.new_int_var(0, n, f"over_off_{m}")
        model.add(sum(free_off) <= off_days + over_off)
        penalties.append(over_off * 500)  # 強力懲罰超休，盡量不超過應休天數

        # ── 硬規則 2：反向班禁止（統一時間軸：history 邊界併入同一組約束，不特判第一天）
        # 允許模式：E, OFF, D  /  N, OFF, E  /  N, OFF, OFF, D
        # 禁止模式（直接或插非 OFF 換班）：
        #   E→D：E 後 1 天不能是 D（需先有 1 天 OFF）
        #   N→E：N 後 1 天不能是 E（需先有 1 天 OFF）
        #   N→D：N 後 1 天或 2 天都不能是 D（需先有 2 天 OFF）
        def _imply_zero(cond, target_var):
            """cond 可能是歷史常數（0/1）或決策變數；target_var 必為決策變數（依呼叫端 range 保證）。"""
            if isinstance(cond, int):
                if cond == 1:
                    model.add(target_var == 0)
            else:
                model.add(target_var == 0).only_enforce_if(cond)

        if no_reverse and "reverse" not in DEBUG_SKIP:
            for t in range(-1, n - 1):
                # E 後緊接 D 禁止（E→D 需要 1 天 OFF 間隔）
                _imply_zero(bx(m, t, 1), b[m][t+1][0])
                # N 後緊接 E 禁止（N→E 需要 1 天 OFF 間隔）
                _imply_zero(bx(m, t, 2), b[m][t+1][1])
                # N 後緊接 D 禁止（N→D 需要 2 天 OFF 間隔）
                _imply_zero(bx(m, t, 2), b[m][t+1][0])
            for t in range(-2, n - 2):
                # N 後第 2 天也不能是 D（1 天 OFF 不夠，要 2 天）
                _imply_zero(bx(m, t, 2), b[m][t+2][0])

        # ── 硬規則 3：每週至少一天應休（OFF 或半），LEAVE_ADJUST（V、員、喪⋯）不計入
        for wi, (ws, we) in enumerate(weeks):
            week_range = list(range(ws, we + 1))
            # 僅計算非 LEAVE_ADJUST 的天（LEAVE_ADJUST 不算應休）
            rest_eligible = [t for t in week_range if t not in leave_adjust_days_m]
            if not rest_eligible:
                continue  # 整週都是 LEAVE_ADJUST，跳過
            # 計算本週「可自由調整」的天（非確認上班且非第一天鎖定上班）
            free_in_week = []
            for t in rest_eligible:
                key = (uid, cycle_dates[t])
                if key not in existing:
                    free_in_week.append(t)
                    continue
                row = existing[key]
                sh  = (row.get("shift") or "OFF")
                locked_work = (
                    (row.get("confirmed") and not overwrite_confirmed and sh not in REST_CODES and sh not in LEAVE_ADJUST)
                    or (lock_first_day and t == 0 and sh not in REST_CODES and sh not in LEAVE_ADJUST)
                )
                if not locked_work:
                    free_in_week.append(t)
            if free_in_week:
                week_rest_sum = sum(b[m][t][3] for t in rest_eligible)
                # S2 每週至少 1 天休（硬規則）
                _a_s2 = model.new_bool_var(f"a_s2_{m}_{wi}")
                model.add(week_rest_sum >= 1).only_enforce_if(_a_s2)
                assume_reg.append((_a_s2, f"每週至少1天休 — {nurse_name} 第{wi+1}週"))
                # S3 一例一休：每週至少 2 天休（硬規則）。湊不出時生成失敗（INFEASIBLE），
                # 由下方失敗診斷提示；不勾此規則則完全不強制。
                if one_in_seven:
                    _a_s3 = model.new_bool_var(f"a_s3_{m}_{wi}")
                    model.add(week_rest_sum >= 2).only_enforce_if(_a_s3)
                    assume_reg.append((_a_s3, f"一例一休(每週≥2天休) — {nurse_name} 第{wi+1}週"))
                # 每週標準休上限 2（軟規則，僅全職）：一例一休真正意思是「正常週剛好 2 天、不多給」。
                # 超過 2 的標準 OFF 會被罰 → 每週收斂為剛好 2。多出的休假只應來自特殊假
                # （V/病/喪等 LEAVE_ADJUST 已排除在 week_rest_sum 外），或國定假日（計入應休總額、
                # 可落在任一週，此時該週容許 3 且僅罰最小值）。半職不套用（應休天數較多，另有邏輯）。
                if not is_ht:
                    # 凸性懲罰（門檻 2/3/4 各罰一次）→ 邊際成本遞增：某週第 3 天 OFF 罰 1 份、
                    # 第 4 天再多罰、第 5 天更多… 逼每週 OFF 平均分布（要 3 都 3，不要某週爆 4/5），
                    # 且盡量壓回每週 2。人力過剩導致的多餘 OFF 無法消除，但會平均攤開。
                    for _th in (2, 3, 4):
                        _ov = model.new_int_var(0, len(rest_eligible), f"wkov_{m}_{wi}_{_th}")
                        model.add(_ov >= week_rest_sum - _th)
                        penalties.append(_ov * WEEKLY_OFF_OVER_PENALTY)
            # 否則整週鎖滿上班 → 跳過約束（異常偵測會在生成後標示警告）

        # ── 硬規則 4：每週 D/E/N 至多兩種班別（絕對硬性，不放寬）
        # 例外日造成的第三種班種，由上方「允許班種限制」放寬自由格來讓路，
        # 使自由格自動收斂為「例外班種 + 其中一種原班別」，維持本規則恆為 <=2
        for ws, we in weeks:
            has_D = model.new_bool_var(f"hasD_{m}_{ws}")
            has_E = model.new_bool_var(f"hasE_{m}_{ws}")
            has_N = model.new_bool_var(f"hasN_{m}_{ws}")
            model.add(sum(b[m][t][0] for t in range(ws, we+1)) >= 1).only_enforce_if(has_D)
            model.add(sum(b[m][t][0] for t in range(ws, we+1)) == 0).only_enforce_if(has_D.negated())
            model.add(sum(b[m][t][1] for t in range(ws, we+1)) >= 1).only_enforce_if(has_E)
            model.add(sum(b[m][t][1] for t in range(ws, we+1)) == 0).only_enforce_if(has_E.negated())
            model.add(sum(b[m][t][2] for t in range(ws, we+1)) >= 1).only_enforce_if(has_N)
            model.add(sum(b[m][t][2] for t in range(ws, we+1)) == 0).only_enforce_if(has_N.negated())
            if "rule4" not in DEBUG_SKIP:
                model.add(has_D + has_E + has_N <= 2)

        # ── 連續上班上限：統一時間軸（history 固定值 + decision 變數）單一滑動視窗，
        # 不特判第一天。視窗起點從 -max_consec 開始，只要視窗觸及決策日就一併檢查；
        # 視窗內的 history 天數固定貢獻直接算進去，決策變數的允許上限相應扣減
        # （若 history 本身已達上限，決策日在該視窗內必須全休）。
        if "consec" not in DEBUG_SKIP:
            for t in range(-max_consec, n - max_consec):
                decision_terms = []
                hist_sum = 0
                for k in range(max_consec + 1):
                    for s in range(3):
                        val = bx(m, t + k, s)
                        if isinstance(val, int):
                            hist_sum += val
                        else:
                            decision_terms.append(val)
                if decision_terms:
                    cap = max(0, max_consec - hist_sum)
                    model.add(sum(decision_terms) <= cap)

        # ── 一例一休（S3）：硬規則（上方 rest 區塊 week_rest_sum >= 2），此處不重複強制

        # ── 首個週末不同時休：此規則已改為「護理師預班階段硬擋」（前端 NursePage 處理），
        #    自動排班不再受此約束（自動排班若給首週末雙休沒有問題）。

        # ── 規則 7：連續 OFF 總天數上限（半職護理師跳過，因應休天數多）
        if not is_ht and "rule7" not in DEBUG_SKIP:
            for t in range(n - weekly_max_off_total):
                if any((t + k) in leave_adjust_days_m for k in range(weekly_max_off_total + 1)):
                    continue
                # 若視窗內已鎖定的 OFF 超過上限（既成事實），跳過此限制
                locked_off_in_win = sum(
                    1 for k in range(weekly_max_off_total + 1)
                    if (t + k) in locked_off_days_m
                )
                if locked_off_in_win > weekly_max_off_total:
                    continue
                model.add(
                    sum(b[m][t + k][3] for k in range(weekly_max_off_total + 1))
                    <= weekly_max_off_total
                )

        # ── 規則 6：自動休連續天數上限（半職護理師跳過）
        if not is_ht and "rule6" not in DEBUG_SKIP:
            for t in range(n - weekly_max_off_auto):
                auto_off_win = [
                    b[m][t+k][3]
                    for k in range(weekly_max_off_auto + 1)
                    if (t+k) not in locked_off_days_m and (t+k) not in leave_adjust_days_m
                ]
                if len(auto_off_win) > weekly_max_off_auto:
                    model.add(sum(auto_off_win) <= weekly_max_off_auto)

        leave_adjust_per_m[m] = leave_adjust_days_m

    # ── 硬規則 1：每班每日剛好 req 人；S11/S12（每班至少 1 leader、至少 2 leader/second）為硬規則，人力不足時會 INFEASIBLE
    leaders = [i for i, n in enumerate(nurses) if n.get("level") == "leader"]
    seconds = [i for i, n in enumerate(nurses) if n.get("level") in ("leader", "second")]

    SHIFT_ALLOWED_MAP = {
        "固定D": ["D"], "固定E": ["E"], "固定N": ["N"],
        "輪班DE": ["D", "E"], "輪班EN": ["E", "N"], "輪班DN": ["D", "N"], "輪班DEN": ["D", "E", "N"],
    }
    _WORK_SHIFTS_LIST = ["D", "E", "N"]
    # 各班別有能力上班的 leader/second 清單（用於檢查可行性）
    _capable_leaders = {
        si: [m for m in leaders if _WORK_SHIFTS_LIST[si] in SHIFT_ALLOWED_MAP.get(nurses[m].get("attr") or "輪班DEN", _WORK_SHIFTS_LIST)]
        for si in range(3)
    }
    _capable_seconds = {
        si: [m for m in seconds if _WORK_SHIFTS_LIST[si] in SHIFT_ALLOWED_MAP.get(nurses[m].get("attr") or "輪班DEN", _WORK_SHIFTS_LIST)]
        for si in range(3)
    }
    # 硬約束（人力不足時會導致 INFEASIBLE，於下方失敗診斷會提示 leader/second 不足）：
    #   S11：每班每日至少 1 位 leader
    #   S12：每班每日至少 2 位 leader/second（受當班需求上限 min(2, req)）

    # 硬規則 1：每班每日剛好 req 人（req==0 同樣加入，防止 solver 自由分配無需求班別）
    # 預填/確認格可能與需求人數矛盾 → 改為高懲罰軟約束，偏差於生成後回報警告
    # S1 每班每日人數：硬規則（恰好符合需求）。若因預填/已確認資料湊不齊，
    # 會生成失敗並由下方失敗診斷（人數超標 / 各班可排人數不足）提示原因。
    demand_dev_vars: list[tuple[int, int, object, object]] = []  # 保留供下游相容（硬規則下為空）
    _snames_diag = ["白班D", "小夜E", "大夜N"]
    if "demand" not in DEBUG_SKIP:
        for t in range(n):
            for si, req in [(0, day_d[t]), (1, day_e[t]), (2, day_n[t])]:
                # 排除行政班格（視同 D 但不佔臨床人力；行政班只落在 D，E/N 不受影響）
                _terms = [b[m][t][si] for m in range(M) if (m, t) not in admin_cells and m not in trainee_set]
                _a_s1 = model.new_bool_var(f"a_s1_{t}_{si}")
                model.add(sum(_terms) == req).only_enforce_if(_a_s1)
                assume_reg.append((_a_s1, f"{_snames_diag[si]}需恰好{req}人（不含行政班）— {cycle_dates[t]}"))

    # ── 軟規則：順班目標 + 固定班偏離懲罰 + leader/second 出勤偏好
    FIX_PENALTY = 500
    # 固定班最多允許偏離的格數（硬上限），由「允許固定班偏離」規則控制：
    #   勾選(allow_fixed_deviation=True)＝最多偏離 2 格；未勾＝0（固定班完全不可偏離）。
    # fair（公平優先版）一律嚴格 0（固定D只排D）；逃生閥 FAIR_FIX_DEV 可放寬 fair。
    _allow_fixed_dev = bool(scheduling.get("allow_fixed_deviation", True))
    FIX_MAX_DEVIATION = int(os.getenv("FAIR_FIX_DEV", "0")) if profile == "fair" else (2 if _allow_fixed_dev else 0)
    # leader/second 軟懲罰（先佔位，懲罰值後加入）
    leader_miss_vars: list[tuple[object, int]] = []   # (bool_var, penalty)
    for t in range(n):
        for si in range(3):
            req = [day_d[t], day_e[t], day_n[t]][si]
            if req == 0:
                continue  # 需求為 0，跳過
            # S11 每班每日至少 1 位 leader（硬規則）；排除行政班（行政班 leader 不算臨床帶班）
            cl = _capable_leaders[si]
            if cl:
                _terms11 = [b[m][t][si] for m in cl if (m, t) not in admin_cells and m not in trainee_set]
                _a_s11 = model.new_bool_var(f"a_s11_{t}_{si}")
                model.add(sum(_terms11) >= 1).only_enforce_if(_a_s11)
                assume_reg.append((_a_s11, f"{_snames_diag[si]}至少1位leader（不含行政班）— {cycle_dates[t]}"))
            # S12 每班每日至少 2 位 leader/second（硬規則，受當班需求上限 min(2, req)）；排除行政班
            cs = _capable_seconds[si]
            if len(cs) >= 2:
                target_s = min(2, req)
                _terms12 = [b[m][t][si] for m in cs if (m, t) not in admin_cells and m not in trainee_set]
                _a_s12 = model.new_bool_var(f"a_s12_{t}_{si}")
                model.add(sum(_terms12) >= target_s).only_enforce_if(_a_s12)
                assume_reg.append((_a_s12, f"{_snames_diag[si]}至少{target_s}位leader/second（不含行政班）— {cycle_dates[t]}"))
    # ── 新人跟隨導師（軟性）：新人每天盡量與導師同班；新人自己的預班/請假已鎖為硬、
    #    每週剛好2休等規則照常。衝突時求解器彈性偏離幾天（用戶確認做法A：新人自己的休算進其2天週休）。
    for _ti, _mj in mentor_of.items():
        for t in range(n):
            _mdiff = model.new_bool_var(f"mentor_diff_{_ti}_{t}")
            model.add(x[_ti][t] != x[_mj][t]).only_enforce_if(_mdiff)
            model.add(x[_ti][t] == x[_mj][t]).only_enforce_if(_mdiff.negated())
            penalties.append(_mdiff * MENTOR_FOLLOW_PENALTY)

    # off_slack 懲罰（高代價，只在人力不足時才縮減）
    for _, slack_var in off_slack_vars:
        penalties.append(slack_var * 200)
    # 軟規則 E：off_slack 公平分配，懲罰最大與最小差距（避免集中同一人）
    if off_slack_vars:
        _max_sk = model.new_int_var(0, 2, "max_off_slack")
        _min_sk = model.new_int_var(0, 2, "min_off_slack")
        for _, sv in off_slack_vars:
            model.add(_max_sk >= sv)
            model.add(_min_sk <= sv)
        _sk_spread = model.new_int_var(0, 2, "slack_spread")
        model.add(_sk_spread == _max_sk - _min_sk)
        penalties.append(_sk_spread * 400 * FAIR_MULT)
    # leader/second 軟約束懲罰
    for miss_var, pen in leader_miss_vars:
        penalties.append(miss_var * pen)
    for m in range(M):
        attr = nurses[m].get("attr") or "輪班DEN"
        fixed_si = FIXED_SHIFT_MAP.get(attr)
        la_set = leave_adjust_per_m.get(m, set())

        # ── 班次比例偏差懲罰（C-2）：班種差值法，±1 天彈性
        # 直接懲罰班種之間的差值，不依賴固定工作天數目標
        # 當 off_slack 被啟用時工作天增加，差值約束仍能維持比例平衡
        # 另設硬上限：全職、半職每班種偏差皆 ≤ ±2 天（預填鎖定已超過時自動讓路）
        DIST_PENALTY = 900 * FAIR_MULT
        _is_ht_m = bool(nurses[m].get("halftime"))
        _cap_days = 2   # 全職、半職皆 ±2（各班種偏離理想 ±2 天）
        _locked_cnt = locked_si_counts_per_m.get(m, [0, 0, 0])
        _total_d  = sum(b[m][t][0] for t in range(n))
        _total_e  = sum(b[m][t][1] for t in range(n))
        _total_nv = sum(b[m][t][2] for t in range(n))
        _uid_m = nurses[m]["uid"]
        _ov_m  = ratio_overrides.get(_uid_m, {})
        def _r(key): return max(1, int(_ov_m.get(key, ratio.get(key, 1))))

        def _add_pair_penalty(va, ra, vb, rb, label, locked_a=0, locked_b=0):
            """懲罰 |va*rb - vb*ra|，tol = ra+rb-1（1:1 時 tol=1）；
            硬上限 |diff| ≤ cap_days*(ra+rb)，若鎖定格已超過則放寬至鎖定值"""
            tol  = ra + rb - 1
            diff = model.new_int_var(-(n * max(ra, rb)), n * max(ra, rb), f"pdiff_{label}")
            dev  = model.new_int_var(0, n * max(ra, rb), f"pdev_{label}")
            model.add(diff == va * rb - vb * ra)
            model.add(dev >= diff - tol)
            model.add(dev >= -diff - tol)
            model.add(dev >= 0)
            penalties.append(dev * DIST_PENALTY)
            # 硬上限（±cap_days 天，換算差值尺度）；鎖定既成事實已超過時讓路
            hard_cap = _cap_days * (ra + rb)
            locked_diff = abs(locked_a * rb - locked_b * ra)
            hard_cap = max(hard_cap, locked_diff)
            model.add(diff <= hard_cap)
            model.add(diff >= -hard_cap)

        if attr == "輪班DE":
            _add_pair_penalty(_total_d, _r("D"), _total_e, _r("E"), f"de_{m}", _locked_cnt[0], _locked_cnt[1])
        elif attr == "輪班DN":
            _add_pair_penalty(_total_d, _r("D"), _total_nv, _r("N"), f"dn_{m}", _locked_cnt[0], _locked_cnt[2])
        elif attr == "輪班EN":
            _add_pair_penalty(_total_e, _r("E"), _total_nv, _r("N"), f"en_{m}", _locked_cnt[1], _locked_cnt[2])
        elif attr == "輪班DEN":
            _add_pair_penalty(_total_d, _r("D"), _total_e, _r("E"), f"de_{m}", _locked_cnt[0], _locked_cnt[1])
            _add_pair_penalty(_total_d, _r("D"), _total_nv, _r("N"), f"dn_{m}", _locked_cnt[0], _locked_cnt[2])
        else:
            pass  # 固定班由 FIX_PENALTY 處理

        # ── 軟規則：孤立上班日懲罰（OFF-上班-OFF，出來上一天班很累）
        ISOLATED_WORK_PENALTY = 750
        for t in range(1, n - 1):
            iso = model.new_bool_var(f"isowork_{m}_{t}")
            work_t = sum(b[m][t][s] for s in range(3))
            # iso = 1 若 前一天OFF + 當天上班 + 後一天OFF
            model.add(iso >= b[m][t-1][3] + work_t + b[m][t+1][3] - 2)
            penalties.append(iso * ISOLATED_WORK_PENALTY)

        if fixed_si is not None:
            # 固定班：偏離固定班種每格罰 FIX_PENALTY；並硬性限制偏離格數上限（fair版=0，其他版=2）
            deviation_terms = []
            for t in range(n):
                if t not in la_set:
                    for s in range(3):
                        if s != fixed_si:
                            penalties.append(b[m][t][s] * FIX_PENALTY)
                            deviation_terms.append(b[m][t][s])
            if deviation_terms:
                # 硬上限：偏離格數 ≤ FIX_MAX_DEVIATION；預填鎖定已超過時放寬至既成值，避免 INFEASIBLE
                locked_dev = sum(_locked_cnt[s] for s in range(3) if s != fixed_si)
                _fix_cap = max(FIX_MAX_DEVIATION, locked_dev)
                _a_fix = model.new_bool_var(f"a_fix_{m}")
                model.add(sum(deviation_terms) <= _fix_cap).only_enforce_if(_a_fix)
                assume_reg.append((_a_fix, f"{attr}固定班別(偏離≤{_fix_cap}格) — {nurse_name}"))
        else:
            # 軟規則 S8：順班。輪班本來就必須換班（2種班最少1次、3種班最少2次），
            # 這些「必要換班」不罰；只罰「多餘換班（超過必要數）」＋「沒休就換」加扣。
            allowed = SHIFT_ALLOWED.get(attr, WORK_SHIFTS)
            num_types = len([c for c in allowed if c in WORK_SHIFTS])
            if num_types <= 1:
                continue
            min_sw = num_types - 1               # 必要換班數（2種班=1、DEN=2、固定班=0）
            max_gap = weekly_max_off_total       # Rule7 限制最大連續 OFF
            E = int(os.getenv("EXCESS_SWITCH_PENALTY", "1500")) * SWITCH_MULT   # 多餘換班扣分（超過必要數的每次）
            R = int(os.getenv("DIRECT_SWITCH_PENALTY", "500")) * SWITCH_MULT    # 沒休就換加扣（直接切換，必要或多餘皆計）

            all_sw_vars = []      # 每一次換班（含直接與隔OFF）
            direct_sw_vars = []   # 直接沒休就換（g=0）
            # 班型轉換偵測（統一時間軸：history 併入，不特判第一天）
            for t in range(n):
                for s2 in range(3):
                    if WORK_SHIFTS[s2] not in allowed:
                        continue
                    max_g = min(max_gap, t + HISTORY_DAYS - 1)
                    for g in range(max_g + 1):  # g = 中間連續 OFF 天數
                        t1 = t - g - 1
                        if t1 < -HISTORY_DAYS:
                            continue
                        for s1 in range(3):
                            if s1 == s2:
                                continue
                            if WORK_SHIFTS[s1] not in allowed:
                                continue
                            parts = ([bx(m, t1, s1)]
                                     + [bx(m, t1 + k, 3) for k in range(1, g + 1)]
                                     + [bx(m, t, s2)])
                            const_sum = sum(p for p in parts if isinstance(p, int))
                            var_parts = [p for p in parts if not isinstance(p, int)]
                            sw = model.new_bool_var(f"gsw_{m}_{t}_{g}_{s1}_{s2}")
                            model.add(sw >= sum(var_parts) + const_sum - (g + 1))
                            all_sw_vars.append(sw)
                            if g == 0:
                                direct_sw_vars.append(sw)
            if all_sw_vars:
                # 多餘換班 = max(0, 總換班 − 必要換班數)，每次扣 E
                excess = model.new_int_var(0, n, f"excess_sw_{m}")
                model.add(excess >= sum(all_sw_vars) - min_sw)
                penalties.append(excess * E)
                # 沒休就換：每次直接切換另加扣 R（必要或多餘皆計）
                if direct_sw_vars:
                    penalties.append(sum(direct_sw_vars) * R)

    model.minimize(sum(penalties))

    # ── 求解前診斷
    print(f"\n[SOLVE] 護理師={M} 週期={n}天 需求D/E/N={daily_d}/{daily_e}/{daily_n}")
    print(f"[SOLVE] 總工作需求={total_work_demand}  總可提供={sum((n - (part_off if nurses[m].get('halftime') else full_off)) for m in range(M))}")
    print(f"[SOLVE] 反向班={no_reverse} 一例一休={one_in_seven} 連班上限={max_consec} full_off={full_off} part_off={part_off}")
    print(f"[SOLVE] 變數數={model.proto.variables.__len__()}  約束數={len(model.proto.constraints)}")
    # 各班可排護理師數
    cap_d = sum(1 for m in range(M) if 0 in {SI.get(s) for s in SHIFT_ALLOWED.get(nurses[m].get("attr","輪班DEN"), WORK_SHIFTS)})
    cap_e = sum(1 for m in range(M) if 1 in {SI.get(s) for s in SHIFT_ALLOWED.get(nurses[m].get("attr","輪班DEN"), WORK_SHIFTS)})
    cap_n = sum(1 for m in range(M) if 2 in {SI.get(s) for s in SHIFT_ALLOWED.get(nurses[m].get("attr","輪班DEN"), WORK_SHIFTS)})
    print(f"[SOLVE] 可排D班護理師={cap_d}人（需求{daily_d}）可排E班={cap_e}人（需求{daily_e}）可排N班={cap_n}人（需求{daily_n}）")
    for m in range(M):
        attr = nurses[m].get("attr","?")
        ht = nurses[m].get("halftime", False)
        allowed = SHIFT_ALLOWED.get(attr, WORK_SHIFTS)
        fixed = FIXED_SHIFT_MAP.get(attr)
        print(f"[NURSE] {nurses[m].get('name','?')} attr={attr} halftime={ht} allowed={list(allowed)} fixed={fixed}")

    # ── 預先衝突診斷（結果回傳給前端）
    pre_conflicts: list[str] = []

    # 診斷 1：確認班 vs 護理師允許班種（已由 exception_days_m 處理，列為警告而非硬錯誤）
    for m in range(M):
        uid = nurses[m]["uid"]
        name = nurses[m].get("name") or uid
        attr_m = nurses[m].get("attr", "輪班DEN")
        allowed_m = SHIFT_ALLOWED.get(attr_m, WORK_SHIFTS)
        allowed_si_m = {SI.get(s) for s in allowed_m} | {3}
        fixed_si_m = FIXED_SHIFT_MAP.get(attr_m)
        for t, d_str in enumerate(cycle_dates):
            key = (uid, d_str)
            if key not in existing:
                continue
            row = existing[key]
            shift = row.get("shift") or "OFF"
            if not row.get("confirmed"):
                continue
            si = SI.get(shift)
            if si is None or si == 3:
                continue
            if fixed_si_m is None and si not in allowed_si_m:
                msg = f"【班種衝突】{name}（{attr_m}）{d_str} 已確認 {shift} 班，不在允許班別 {'/'.join(allowed_m)}（已保留，CP-SAT 將於後續導正）"
                prefill_conflicts.append(msg)
                print(f"[CONFLICT] {msg}")

    # 診斷 2：確認班人數 > 當日需求
    locked_by_day: dict[int, dict[int, int]] = {}
    for m in range(M):
        uid = nurses[m]["uid"]
        for t, d_str in enumerate(cycle_dates):
            key = (uid, d_str)
            if key not in existing:
                continue
            row = existing[key]
            shift = row.get("shift") or "OFF"
            if not row.get("confirmed"):
                continue
            si = SI.get(shift)
            if si is None or si == 3:
                continue
            locked_by_day.setdefault(t, {0:0, 1:0, 2:0})[si] += 1
    for t, counts in locked_by_day.items():
        for si, req, sname in [(0, day_d[t], "D"), (1, day_e[t], "E"), (2, day_n[t], "N")]:
            if counts[si] > req:
                msg = f"【人數超標】{cycle_dates[t]} {sname}班已確認 {counts[si]} 人，超過需求 {req} 人"
                pre_conflicts.append(msg)
                print(f"[CONFLICT] {msg}")

    # 診斷 3：同一週已鎖定（含未確認之預填）班種超過 2 種 → 與「每週至多兩種班別」硬規則
    # 恆無法並存的真衝突（此規則不放寬，須靠人工調整預填資料）
    for m in range(M):
        uid = nurses[m]["uid"]
        name = nurses[m].get("name") or uid
        attr_m = nurses[m].get("attr", "輪班DEN")
        if attr_m == "輪班DEN":
            continue  # DEN 本身允許三種班，不受此限
        for ws, we in weeks:
            week_si: dict[int, list[str]] = {}
            for t in range(ws, we + 1):
                d_str = cycle_dates[t]
                row = existing.get((uid, d_str))
                if not row or not row.get("shift"):
                    continue
                si = SI.get(row["shift"])
                if si is None or si == 3:
                    continue
                week_si.setdefault(si, []).append(d_str)
            if len(week_si) > 2:
                detail_days = "、".join(f"{d}={['D','E','N'][si]}" for si, ds in week_si.items() for d in ds)
                msg = f"【每週超過兩種班】{name}（{attr_m}）{cycle_dates[ws]}~{cycle_dates[we]} 已有 {len(week_si)} 種班別預填：{detail_days}，違反「每週至多兩種班別」硬規則，請調整其中一格"
                pre_conflicts.append(msg)
                print(f"[CONFLICT] {msg}")

    # 注意：prefill_conflicts 是警告（已由例外日機制處理），不列入 INFEASIBLE 原因

    # 掛上 assumption 開關（全部設為真＝硬規則生效；presolve 會化簡掉，正常求解幾乎無額外開銷）。
    # 一旦無解，可用 solver.sufficient_assumptions_for_infeasibility() 反查互相矛盾的規則。
    if assume_reg:
        model.add_assumptions([lit for lit, _ in assume_reg])

    solver = cp_model.CpSolver()
    solver.parameters.max_time_in_seconds = 90
    solver.parameters.num_workers = 4
    status = solver.solve(model)
    print(f"[SOLVE] status={solver.status_name(status)}  wall_time={solver.wall_time:.1f}s")

    # ── 主解失敗（UNKNOWN 逾時 或 INFEASIBLE）：跑一次「純可行性」診斷解。
    #    移除目標函數（最佳化才是主要負擔）＋較短時限，讓求解器能真正判定可行性：
    #    可行 → 把此可行解救回來當結果（未最佳化）；不可行 → 供下方取 infeasible core 點名衝突。
    diag_solver = None
    diag_status = None
    rescued_warning = None
    if status not in (cp_model.OPTIMAL, cp_model.FEASIBLE) and assume_reg:
        try:
            model.proto.ClearField("objective")
            diag_solver = cp_model.CpSolver()
            diag_solver.parameters.max_time_in_seconds = float(os.getenv("DIAG_SOLVE_SECONDS", "45"))
            diag_solver.parameters.num_workers = 8
            diag_status = diag_solver.solve(model)
            print(f"[SOLVE-DIAG] status={diag_solver.status_name(diag_status)}  wall_time={diag_solver.wall_time:.1f}s")
        except Exception as _e:
            print(f"[SOLVE-DIAG] 診斷解失敗: {_e}")
        if diag_status in (cp_model.OPTIMAL, cp_model.FEASIBLE):
            solver = diag_solver
            status = diag_status
            rescued_warning = ("此版在時限內無法完成最佳化，改用「可行但未最佳化」的班表："
                               "所有硬規則（含固定班／一例一休等）皆已滿足，但順班、公平、班次比例等軟目標未經優化，"
                               "數據可能較差。可改選其他版本，或減少已確認/預班衝突後重試。")
            print("[SOLVE-DIAG] 採用診斷可行解（救援）")

    if status not in (cp_model.OPTIMAL, cp_model.FEASIBLE):
        violations: list[str] = []

        # 0. Infeasible core：求解器回報「實際互相矛盾的硬規則」（最精準，直接點名規則+護理師+週次/日期）。
        #    優先用診斷解證明的無解（較可靠）；診斷解也逾時則視為模型過難、非明確衝突。
        core_labels: list[str] = []
        _core_solver = (diag_solver if diag_status == cp_model.INFEASIBLE
                        else solver if status == cp_model.INFEASIBLE else None)
        if _core_solver is not None and assume_reg:
            try:
                idx_map = {lit.index: lbl for lit, lbl in assume_reg}
                seen: set[str] = set()
                for i in _core_solver.sufficient_assumptions_for_infeasibility():
                    lbl = idx_map.get(i) or idx_map.get(abs(i))
                    if lbl and lbl not in seen:
                        seen.add(lbl)
                        core_labels.append(lbl)
            except Exception as _e:
                print(f"[SOLVE] core 解析失敗: {_e}")
        if not core_labels and diag_status is not None and diag_status not in (
            cp_model.INFEASIBLE, cp_model.OPTIMAL, cp_model.FEASIBLE
        ):
            violations.append("求解器在時限內無法判定此組合是否可行（模型過大或過難），並非明確的規則衝突；"
                              "建議減少人力變動、縮短週期，或放寬部分軟規則後重試")
        if core_labels:
            shown = core_labels[:12]
            more = f"（另有 {len(core_labels) - 12} 項同類）" if len(core_labels) > 12 else ""
            violations.append("以下規則同時成立時互相矛盾（無法全部滿足）：" + "；".join(shown) + more)

        # 1. 護理師人數 vs 每日總需求
        max_daily_demand = max(day_d[t] + day_e[t] + day_n[t] for t in range(n))
        if M < max_daily_demand:
            violations.append(f"護理師總數（{M}人）少於單日最高需求（{max_daily_demand}人）")

        # 2. 每位護理師可排工作天數 vs 總工作需求
        # 可排工作天數上限 = n - 最少休假天數
        min_off_per_nurse = max(1, n // 7)  # 每週至少 1 天休
        max_work_per_nurse = n - min_off_per_nurse
        max_total_work = M * max_work_per_nurse
        if total_work_demand > max_total_work:
            violations.append(
                f"總工作需求（{total_work_demand}班次）超過所有護理師可排上限"
                f"（{M}人 × {max_work_per_nurse}天 = {max_total_work}班次）"
            )

        # 3. 每日各班需求 vs 可排該班護理師數
        def capable_for(si: int) -> int:
            return sum(
                1 for m in range(M)
                if si in {SI.get(s) for s in SHIFT_ALLOWED.get(nurses[m].get("attr", "輪班DEN"), WORK_SHIFTS)}
            )
        capable_d = capable_for(0)
        capable_e = capable_for(1)
        capable_n = capable_for(2)
        for t in range(n):
            if day_d[t] > capable_d:
                violations.append(f"{cycle_dates[t]}：D班需要{day_d[t]}人但只有{capable_d}人可排D班")
                break
            if day_e[t] > capable_e:
                violations.append(f"{cycle_dates[t]}：E班需要{day_e[t]}人但只有{capable_e}人可排E班")
                break
            if day_n[t] > capable_n:
                violations.append(f"{cycle_dates[t]}：N班需要{day_n[t]}人但只有{capable_n}人可排N班")
                break

        # 4. 連班上限是否太嚴
        if max_consec > 0 and max_consec < 3:
            violations.append(f"連續上班上限（{max_consec}天）過低，難以滿足每日需求")

        # 5. 連休上限
        max_consec_off_val = scheduling.get("max_consec_off", 3)
        if max_consec_off_val < 1:
            violations.append("連續休假上限不可為 0")

        # 5b. 一例一休（硬規則）：每人每週至少 2 天休 → 可排天數上限下降，可能導致人力不足
        if one_in_seven:
            num_weeks = max(1, len(weeks))
            max_work_2off = max(0, n - 2 * num_weeks)
            if total_work_demand > M * max_work_2off:
                violations.append(
                    f"「一例一休」為硬規則（每人每週至少 2 天休）使可排天數下降："
                    f"總工作需求（{total_work_demand}班次）超過可排上限"
                    f"（{M}人 × {max_work_2off}天 = {M * max_work_2off}班次）。"
                    f"建議取消勾選一例一休、增加人力或降低各班需求人數"
                )

        # 6. leader/second 人力是否足以滿足 S11/S12 硬規則
        _snames = ["白班D", "小夜E", "大夜N"]
        for si in range(3):
            has_demand = any([day_d, day_e, day_n][si][t] > 0 for t in range(n))
            if not has_demand:
                continue
            n_leader = len(_capable_leaders[si])
            n_second = len(_capable_seconds[si])
            if n_leader == 0:
                violations.append(f"{_snames[si]}沒有任何可上此班的 leader，無法滿足『每班至少 1 位 leader』硬規則（請指派 leader 或放寬此規則）")
            elif n_leader == 1:
                violations.append(f"{_snames[si]}只有 1 位可上此班的 leader，硬性要求他每天在班，扣掉輪休後幾乎不可能（建議增加 leader 或放寬）")
            if n_second < 2:
                violations.append(f"{_snames[si]}可上此班的 leader/second 不足 2 位（目前 {n_second} 位），無法滿足『每班至少 2 位 leader/second』硬規則")

        # 優先顯示預先偵測到的具體衝突
        if pre_conflicts:
            detail = "⚠ 無法生成班表，發現以下衝突：\n" + "\n".join(pre_conflicts)
        elif violations:
            detail = "⚠ 無法生成符合所有規則的班表。診斷結果：" + "；".join(violations)
        else:
            detail = "⚠ 無法生成班表。規則條件衝突，建議：① 增加護理師人數 ② 降低每日各班需求人數 ③ 放寬連班/反向班限制"
            if one_in_seven:
                detail += " ④ 若某週人力吃緊，可取消勾選「一例一休（每週≥2天休）」（現為硬規則）"
        if prefill_conflicts:
            # 這些預班/確認資料與規則不符，在嚴格規則（如公平版固定班）下常是排不出的主因 → 提為「建議優先修正」
            detail += "\n\n⚠ 以下預班／已確認資料與規則不符，嚴格規則下常是排不出的主因，建議先修正：\n" + "\n".join(prefill_conflicts)

        raise HTTPException(400, detail)

    # ── 解析結果
    SHIFT_NAMES = ["D", "E", "N", "OFF"]
    schedules: dict[str, list[str]] = {}
    for m, nurse in enumerate(nurses):
        sched = [SHIFT_NAMES[solver.value(x[m][t])] for t in range(n)]
        # 還原特休、指定休（不被 CP-SAT 覆蓋，從 existing 讀回）
        for t, d_str in enumerate(cycle_dates):
            key = (nurse["uid"], d_str)
            if key in existing:
                orig = existing[key].get("shift") or "OFF"
                if orig in LEAVE_ADJUST:
                    sched[t] = orig  # 放假/調整類：保留原班碼
                elif orig in ADMIN_SHIFTS:
                    sched[t] = orig  # 行政班：保留原班碼
                elif orig in REST_CODES and orig != "OFF":
                    sched[t] = orig  # 應休類（如半）：保留原班碼
        schedules[nurse["uid"]] = sched

    # ── 人力不足警告（off_slack > 0）
    warnings: list[str] = []
    if rescued_warning:
        warnings.append("⚠ " + rescued_warning)
    reduced_nurses = []
    for mi, slack_var in off_slack_vars:
        v = solver.value(slack_var)
        if v > 0:
            nm = nurses[mi].get("name") or nurses[mi]["uid"]
            reduced_nurses.append(f"{nm}（減 {v} 天）")
    if reduced_nurses:
        warnings.append("⚠ 人力不足，以下護理師應休天數已自動縮減：" + "、".join(reduced_nurses))

    # ── 異常偵測
    anomalies: list[str] = []
    SHIFT_NAMES_DETECT = ["D", "E", "N", "OFF"]
    for mi, nurse in enumerate(nurses):
        sched_m = [SHIFT_NAMES_DETECT[solver.value(x[mi][t])] for t in range(n)]
        nm = nurse.get("name") or nurse["uid"]
        prev_m = hist_raw_by_nurse.get(nurse["uid"], ["OFF"] * HISTORY_DAYS)
        combined = prev_m + sched_m
        consec = 0
        for ps in combined:
            if ps not in REST_CODES and ps not in LEAVE_ADJUST and ps != "OFF":
                consec += 1
                if consec > max_consec:
                    anomalies.append(f"⚠ {nm}：跨週連續上班超過 {max_consec} 天（含上週）")
                    break
            else:
                consec = 0
    leader_indices = [i for i, n in enumerate(nurses) if n.get("level") == "leader"]
    for t in range(n):
        for si, req, sh in [(0, day_d[t], "D"), (1, day_e[t], "E"), (2, day_n[t], "N")]:
            if req == 0:
                continue
            if not any(solver.value(b[li][t][si]) for li in leader_indices):
                anomalies.append(f"⚠ {cycle_dates[t]} {sh}班：無 leader 排班")

    # ── Post-solve 驗證：確認每日人數剛好等於需求
    demand_violations: list[str] = []
    for t in range(n):
        actual_d = sum(solver.value(b[m][t][0]) for m in range(M))
        actual_e = sum(solver.value(b[m][t][1]) for m in range(M))
        actual_n = sum(solver.value(b[m][t][2]) for m in range(M))
        if actual_d != day_d[t] or actual_e != day_e[t] or actual_n != day_n[t]:
            demand_violations.append(
                f"⚠ {cycle_dates[t]}：D={actual_d}（需{day_d[t]}）E={actual_e}（需{day_e[t]}）N={actual_n}（需{day_n[t]}）"
            )
    if demand_violations:
        for v in demand_violations:
            anomalies.append(v)

    # ── 計算格子數（供前端顯示）
    existing_map_keys: set[str] = {
        f"{r['nurse_uid']}_{r['date']}" for r in (existing_res.data or [])
    }
    existing_confirmed: set[str] = {
        f"{r['nurse_uid']}_{r['date']}"
        for r in (existing_res.data or []) if r.get("confirmed")
    }
    new_cells, update_cells = 0, 0
    for nurse_uid, sched in schedules.items():
        for i, shift in enumerate(sched):
            key = f"{nurse_uid}_{cycle_dates[i]}"
            if key in existing_map_keys:
                if overwrite_confirmed or key not in existing_confirmed:
                    update_cells += 1
            else:
                new_cells += 1

    # ── schedules 轉為 {nurse_uid: {date: shift}} 方便前端傳回 commit（含 OFF）
    schedules_dict = {
        uid: {cycle_dates[i]: s for i, s in enumerate(sched)}
        for uid, sched in schedules.items()
    }

    # ── 需求人數偏差警告（預填/確認格造成無法剛好滿足需求）
    _SNAMES = ["D", "E", "N"]
    for t, si, short, over in demand_dev_vars:
        sv, ov = solver.value(short), solver.value(over)
        if sv > 0 or ov > 0:
            req = [day_d[t], day_e[t], day_n[t]][si]
            actual = req - sv + ov
            prefill_conflicts.append(
                f"【人數偏差】{cycle_dates[t]} {_SNAMES[si]}班排定 {actual} 人（需求 {req} 人），受預填/已確認班表限制"
            )

    # ── 版本比較指標：多餘換班（扣掉必要換班）、孤立上班日、最大比例偏差
    _work_set = {"D", "E", "N"}
    _attr_of = {nz["uid"]: (nz.get("attr") or "輪班DEN") for nz in nurses}
    metric_switches = 0   # 總換班（保留供參考）
    metric_excess = 0     # 多餘換班 = 總換班 − 必要換班數（每人加總）
    metric_isolated = 0
    metric_max_dev = 0
    for uid, sched in schedules.items():
        work_seq = [s for s in sched if s in _work_set]
        sw = sum(1 for a2, b2 in zip(work_seq, work_seq[1:]) if a2 != b2)
        metric_switches += sw
        _kinds = SHIFT_ALLOWED.get(_attr_of.get(uid, "輪班DEN"), list(_work_set))
        _ntypes = len([c for c in _kinds if c in _work_set])
        metric_excess += max(0, sw - max(0, _ntypes - 1))   # 扣掉必要換班數
        for i in range(1, len(sched) - 1):
            if sched[i] in _work_set and sched[i-1] not in _work_set and sched[i+1] not in _work_set:
                metric_isolated += 1
    for m2, nurse2 in enumerate(nurses):
        attr2 = nurse2.get("attr") or "輪班DEN"
        if not attr2.startswith("輪班") or attr2 == "輪班":
            continue
        sched2 = schedules.get(nurse2["uid"], [])
        cnt = {s: sched2.count(s) for s in ("D", "E", "N")}
        kinds = SHIFT_ALLOWED.get(attr2, [])
        if len(kinds) >= 2:
            vals = [cnt.get(k, 0) for k in kinds]
            metric_max_dev = max(metric_max_dev, max(vals) - min(vals))

    # ── 診斷 log：確認 OFF 數量
    total_off_generated = sum(
        1 for sched in schedules.values() for s in sched if s == "OFF"
    )
    print(f"\n[GENERATE] daily_d={daily_d} daily_e={daily_e} daily_n={daily_n}  nurses={M}  cycle={n}d")
    print(f"[GENERATE] OFF slots in schedules_dict: {total_off_generated}")
    for uid, sched in list(schedules.items())[:3]:  # 印前3位
        off_days_count = sched.count("OFF")
        print(f"[GENERATE]   {uid}: OFF={off_days_count}  sample={sched[:7]}")

    return {
        "message": f"✓ CP-SAT 計算完成（{len(nurses)} 位護理師，新增 {new_cells} 格、更新 {update_cells} 格）",
        "schedules": schedules_dict,
        "cycle_dates": cycle_dates,
        "overwrite_confirmed": overwrite_confirmed,
        "solver_status": solver.status_name(status),
        "nurses": len(nurses), "new_cells": new_cells, "update_cells": update_cells,
        "warnings": warnings,
        "anomalies": anomalies,
        "prefill_warnings": prefill_conflicts,
        "demand_config": {
            "daily_d": daily_d, "daily_e": daily_e, "daily_n": daily_n,
            "special_dates_count": len(special_dates_map),
            "total_work_demand": total_work_demand,
        },
        "profile": profile,
        "metrics": {
            "switches": metric_switches,
            "excess_switches": metric_excess,
            "isolated_days": metric_isolated,
            "max_ratio_dev": metric_max_dev,
        },
    }


class CommitScheduleBody(BaseModel):
    schedules: dict[str, dict[str, str]]   # {nurse_uid: {date: shift}}
    cycle_dates: list[str]
    overwrite_confirmed: bool = False


@app.post("/schedule/commit")
def commit_schedule(
    body: CommitScheduleBody,
    current_user: dict = Depends(require_roles("admin", "superadmin", "dual")),
):
    """將 generate 回傳的 schedules 寫入資料庫"""
    operator_uid = current_user.get("sub")

    # 讀取規則以取得班別定義（用於分類）
    rules_res = supabase.table("rules").select("*").limit(1).execute()
    rules = rules_res.data[0].get("data") or {} if rules_res.data else {}
    shift_defs = rules.get("shifts", {})
    _rest_defs  = shift_defs.get("rest", [])
    _leave_defs = shift_defs.get("off",  [])
    REST_CODES   = {s["code"] for s in _rest_defs  if s.get("code")} or {"OFF", "半"}
    LEAVE_ADJUST = {s["code"] for s in _leave_defs if s.get("code")} or {"V", "員", "喪", "延休", "補休", "調移"}

    all_dates = body.cycle_dates
    if not all_dates:
        raise HTTPException(400, "無排班日期")

    # 讀取現有班別
    start_str, end_str = all_dates[0], all_dates[-1]
    existing_res = supabase.table("shifts").select("nurse_uid, date, shift, confirmed") \
        .gte("date", start_str).lte("date", end_str).execute()
    existing_map: dict[str, bool] = {
        f"{r['nurse_uid']}_{r['date']}": r.get("confirmed", False)
        for r in (existing_res.data or [])
    }

    to_insert, to_update = [], []
    generated_keys = []
    skipped_confirmed = 0

    for nurse_uid, date_shifts in body.schedules.items():
        for d_str, shift in date_shifts.items():
            key = f"{nurse_uid}_{d_str}"
            if key in existing_map:
                if not body.overwrite_confirmed and existing_map[key]:
                    skipped_confirmed += 1
                    continue  # 已確認且不覆蓋 → 跳過
                to_update.append({"nurse_uid": nurse_uid, "date": d_str, "shift": shift})
            else:
                to_insert.append({
                    "code": key, "label": shift,
                    "nurse_uid": nurse_uid, "date": d_str,
                    "shift": shift, "confirmed": False,
                    "updated_by": operator_uid,
                })
                generated_keys.append(key)

    off_insert = sum(1 for r in to_insert if r["shift"] == "OFF")
    off_update = sum(1 for r in to_update if r["shift"] == "OFF")
    print(f"\n[COMMIT] received={sum(len(v) for v in body.schedules.values())} shifts  overwrite={body.overwrite_confirmed}")
    print(f"[COMMIT] to_insert={len(to_insert)} (OFF={off_insert})  to_update={len(to_update)} (OFF={off_update})  skipped_confirmed={skipped_confirmed}")

    if to_insert:
        supabase.table("shifts").insert(to_insert).execute()
    print(f"[COMMIT] insert done  update count={len(to_update)}")
    for row in to_update:
        supabase.table("shifts").update({
            "shift": row["shift"], "confirmed": False,
            "updated_by": operator_uid,
            "updated_at": datetime.utcnow().isoformat(),
        }).eq("nurse_uid", row["nurse_uid"]).eq("date", row["date"]).execute()

    # 儲存生成鍵（供 Excel 匯出區分人工／系統，以及「清除CP-SAT生成內容」使用）
    rules_res2 = supabase.table("rules").select("id", "data").limit(1).execute()
    if rules_res2.data:
        cur_data = rules_res2.data[0].get("data") or {}
        cur_data["last_generated_keys"] = generated_keys
        cur_data["last_generated_at"] = datetime.utcnow().isoformat()
        cur_data["last_generated_range"] = {"start_date": start_str, "end_date": end_str}
        # 完整快照（供「恢復到上次CP-SAT生成的內容」使用）：本次生成的完整班表
        cur_data["last_generated_full"] = body.schedules
        supabase.table("rules").update({"data": cur_data}).eq("id", rules_res2.data[0]["id"]).execute()

    total = len(to_insert) + len(to_update)
    return {
        "message": f"✓ 已匯入 {total} 格（新增 {len(to_insert)}、更新 {len(to_update)}）",
        "inserted": len(to_insert), "updated": len(to_update),
    }


def _date_range(start_str: str, end_str: str) -> list[str]:
    s = date_type.fromisoformat(start_str)
    e = date_type.fromisoformat(end_str)
    return [(s + timedelta(days=i)).isoformat() for i in range((e - s).days + 1)]


def _make_border(thick=False):
    s = Side(style="medium" if thick else "thin", color="000000")
    return Border(left=s, right=s, top=s, bottom=s)


def _compute_demand(rules: dict, cycle_dates: list[str]) -> tuple[dict[str, tuple[int, int, int]], set[str]]:
    """回傳 {date: (需求D, 需求E, 需求N)} 與特殊日期覆蓋的日期集合。"""
    scheduling = rules.get("scheduling", {})
    daily_d = int(scheduling.get("daily_d", 3))
    daily_e = int(scheduling.get("daily_e", 3))
    daily_n = int(scheduling.get("daily_n", 3))
    special_raw = scheduling.get("special_dates", []) or []
    special_map = {
        sd["date"]: (int(sd.get("d", daily_d)), int(sd.get("e", daily_e)), int(sd.get("n", daily_n)))
        for sd in special_raw if sd.get("date")
    }
    demand = {d: special_map.get(d, (daily_d, daily_e, daily_n)) for d in cycle_dates}
    special_cols = {d for d in cycle_dates if d in special_map}
    return demand, special_cols


def _build_matrix_excel(
    title: str,
    cycle_dates: list[str],
    nurse_rows: list[dict],       # [{uid, name, shifts: {date: display_shift}}] 依顯示順序
    manual_keys: set[str],        # f"{uid}_{date}" → 人工預填（淡黃底）
    rest_codes: set[str] | None = None,          # 計入應休天數的班別（OFF、半）
    demand: dict[str, tuple[int, int, int]] | None = None,  # date → (需求D, 需求E, 需求N)
    special_date_cols: set[str] | None = None,   # 特殊日期覆蓋 → 表頭標黃
) -> io.BytesIO:
    """矩陣式班表：列=護理師、欄=日期；OFF 紅字；人工預填淡黃底；底部 D/E/N 每日統計；
    右側 OFF 天數合計；人數不足當日統計標黃；特殊日期表頭標黃。"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = title

    rest_codes = rest_codes or {"OFF", "半"}
    special_date_cols = special_date_cols or set()

    thin = _make_border(False)
    center = Alignment(horizontal="center", vertical="center")
    fill_header   = PatternFill("solid", fgColor="D9EAF7")
    fill_weekend  = PatternFill("solid", fgColor="EDEDED")
    fill_manual   = PatternFill("solid", fgColor="FFF9DB")   # 很淡的黃：人工預填
    fill_special  = PatternFill("solid", fgColor="FFF176")   # 特殊日期表頭：黃
    fill_short    = PatternFill("solid", fgColor="FFF176")   # 人數不足：黃
    red_font  = Font(color="FF0000", size=10)
    bold_font = Font(bold=True, size=10)
    norm_font = Font(size=10)

    ND = len(cycle_dates)
    date_objs = [date_type.fromisoformat(d) for d in cycle_dates]
    weekend_cols = {ci for ci, d in enumerate(date_objs, 2) if d.weekday() >= 5}
    special_cols = {ci for ci, d in enumerate(cycle_dates, 2) if d in special_date_cols}

    # 週切分（週一~週日，夾到週期內），用於每週 OFF 天數
    weekdays = [d.weekday() for d in date_objs]
    weeks: list[tuple[int, int]] = []
    _i = 0
    while _i < ND:
        wstart = _i - weekdays[_i]
        wend = wstart + 6
        weeks.append((max(0, wstart), min(ND - 1, wend)))
        _i = wend + 1
    NW = len(weeks)

    # 右側統計欄位位置
    d_col = 2 + ND
    e_col = 3 + ND
    n_col = 4 + ND
    week_cols = [5 + ND + w for w in range(NW)]   # 每週 OFF 欄
    last_col = 4 + ND + NW

    # 資深護理師（leader / second）
    senior_uids = {nr["uid"] for nr in nurse_rows if nr.get("level") in ("leader", "second")}

    # Row1：日期（數字）、Row2：星期
    dow_zh = ["一", "二", "三", "四", "五", "六", "日"]
    ws.cell(row=1, column=1, value="姓名").font = bold_font
    ws.cell(row=1, column=1).fill = fill_header
    ws.cell(row=1, column=1).border = thin
    ws.cell(row=1, column=1).alignment = center
    ws.cell(row=2, column=1, value="").border = thin
    for ci, d in enumerate(date_objs, 2):
        c1 = ws.cell(row=1, column=ci, value=d.day)
        c2 = ws.cell(row=2, column=ci, value=dow_zh[d.weekday()])
        for c in (c1, c2):
            c.font = bold_font
            c.alignment = center
            c.border = thin
            c.fill = fill_special if ci in special_cols else (fill_weekend if ci in weekend_cols else fill_header)
    # 右側統計欄表頭
    stat_headers = [(d_col, "D"), (e_col, "E"), (n_col, "N")]
    for w in range(NW):
        stat_headers.append((week_cols[w], f"{w+1}週OFF"))
    for col, label in stat_headers:
        h = ws.cell(row=1, column=col, value=label)
        ws.cell(row=2, column=col, value="")
        for rr in (1, 2):
            c = ws.cell(row=rr, column=col)
            c.font = bold_font; c.alignment = center; c.border = thin; c.fill = fill_header

    # 護理師列
    ri = 3
    for nr in nurse_rows:
        name_cell = ws.cell(row=ri, column=1, value=nr["name"])
        name_cell.font = norm_font
        name_cell.alignment = center
        name_cell.border = thin
        dcnt = ecnt = ncnt = 0
        for ci, d_str in enumerate(cycle_dates, 2):
            shift = nr["shifts"].get(d_str, "")
            if shift == "D": dcnt += 1
            elif shift == "E": ecnt += 1
            elif shift == "N": ncnt += 1
            cell = ws.cell(row=ri, column=ci, value=shift)
            cell.alignment = center
            cell.border = thin
            cell.font = red_font if shift == "OFF" else norm_font
            if f"{nr['uid']}_{d_str}" in manual_keys and shift:
                cell.fill = fill_manual
            elif ci in special_cols:
                cell.fill = fill_special
            elif ci in weekend_cols:
                cell.fill = fill_weekend
        # 右側：D/E/N 天數
        for col, val in ((d_col, dcnt), (e_col, ecnt), (n_col, ncnt)):
            c = ws.cell(row=ri, column=col, value=val)
            c.alignment = center; c.border = thin; c.font = bold_font
        # 右側：每週 OFF 天數
        for w, (a, b) in enumerate(weeks):
            woff = sum(1 for k in range(a, b + 1) if nr["shifts"].get(cycle_dates[k]) in rest_codes)
            c = ws.cell(row=ri, column=week_cols[w], value=woff)
            c.alignment = center; c.border = thin; c.font = bold_font
        ri += 1

    # ── 底部區塊①：各班 leader+second 人數（每日）
    ri += 1
    for sname in ("D", "E", "N"):
        label = ws.cell(row=ri, column=1, value=f"{sname} 資深")
        label.font = bold_font; label.alignment = center; label.border = thin; label.fill = fill_header
        for ci, d_str in enumerate(cycle_dates, 2):
            cnt = sum(1 for nr in nurse_rows if nr["uid"] in senior_uids and nr["shifts"].get(d_str) == sname)
            c = ws.cell(row=ri, column=ci, value=cnt)
            c.alignment = center; c.border = thin; c.font = norm_font
            if ci in weekend_cols: c.fill = fill_weekend
        ri += 1

    # ── 底部區塊②：各班總人數（每日，人數不足標黃）
    for si, sname in enumerate(("D", "E", "N")):
        label = ws.cell(row=ri, column=1, value=f"{sname} 總數")
        label.font = bold_font; label.alignment = center; label.border = thin
        for ci, d_str in enumerate(cycle_dates, 2):
            cnt = sum(1 for nr in nurse_rows if nr["shifts"].get(d_str) == sname)
            cell = ws.cell(row=ri, column=ci, value=cnt)
            cell.alignment = center; cell.border = thin; cell.font = norm_font
            req = demand.get(d_str, (0, 0, 0))[si] if demand else None
            if req is not None and cnt < req:
                cell.fill = fill_short
            elif ci in weekend_cols:
                cell.fill = fill_weekend
        ri += 1

    # 欄寬
    ws.column_dimensions["A"].width = 12
    for ci in range(2, 2 + ND):
        ws.column_dimensions[openpyxl.utils.get_column_letter(ci)].width = 5.5
    for col in (d_col, e_col, n_col):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 6
    for col in week_cols:
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 8
    ws.freeze_panes = "B3"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


@app.get("/export/preview")
def export_preview(
    current_user: dict = Depends(require_roles("admin", "superadmin", "dual")),
):
    """匯出預假狀態：目前所有護理師已填寫的班別（不分系統/人工）"""
    rules_res = supabase.table("rules").select("*").limit(1).execute()
    rules = rules_res.data[0].get("data") or {} if rules_res.data else {}
    cycle = rules.get("cycle", {})
    start_str, end_str = cycle.get("start_date"), cycle.get("end_date")
    if not start_str or not end_str:
        raise HTTPException(400, "請先設定排班週期")
    shift_defs = rules.get("shifts", {})
    rest_codes = {s["code"] for s in shift_defs.get("rest", []) if s.get("code")} or {"OFF", "半"}

    users_res = supabase.table("users").select("uid, name, level").in_("role", ["nurse", "dual"]).order("sort_order").execute()

    shifts_res = supabase.table("shifts").select("nurse_uid, date, shift") \
        .gte("date", start_str).lte("date", end_str).order("date").execute()

    cycle_dates = _date_range(start_str, end_str)
    shift_map: dict[str, dict[str, str]] = {}
    manual_keys: set[str] = set()
    for r in (shifts_res.data or []):
        if not r.get("shift"):
            continue
        shift_map.setdefault(r["nurse_uid"], {})[r["date"]] = r["shift"]
        manual_keys.add(f"{r['nurse_uid']}_{r['date']}")  # 預假匯出：全部都是人工填寫
    nurse_rows = [
        {"uid": u["uid"], "name": u["name"], "level": u.get("level"), "shifts": shift_map.get(u["uid"], {})}
        for u in (users_res.data or [])
    ]

    demand, special_cols = _compute_demand(rules, cycle_dates)
    buf = _build_matrix_excel("預假狀態", cycle_dates, nurse_rows, manual_keys,
                              rest_codes=rest_codes, demand=demand, special_date_cols=special_cols)
    filename = f"預假狀態_{start_str}_{end_str}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f"attachment; filename*=UTF-8''{quote(filename, safe='')}",
            "Cache-Control": "no-store, no-cache, must-revalidate",
        },
    )


@app.get("/export/schedule")
def export_schedule(
    current_user: dict = Depends(require_roles("admin", "superadmin", "dual")),
):
    """匯出完整班表：生成前人工填寫的格子外框加粗；半職轉顯示為休假"""
    rules_res = supabase.table("rules").select("*").limit(1).execute()
    rules = rules_res.data[0].get("data") or {} if rules_res.data else {}
    cycle = rules.get("cycle", {})
    start_str, end_str = cycle.get("start_date"), cycle.get("end_date")
    if not start_str or not end_str:
        raise HTTPException(400, "請先設定排班週期")
    shift_defs = rules.get("shifts", {})
    rest_codes = {s["code"] for s in shift_defs.get("rest", []) if s.get("code")} or {"OFF", "半"}
    # 人工填寫的格子 = 不在 last_generated_keys 裡的格子
    generated_keys = set(rules.get("last_generated_keys", []))

    users_res = supabase.table("users").select("uid, name, level").in_("role", ["nurse", "dual"]).order("sort_order").execute()

    shifts_res = supabase.table("shifts").select("nurse_uid, date, shift") \
        .gte("date", start_str).lte("date", end_str).order("date").execute()

    cycle_dates = _date_range(start_str, end_str)
    shift_map: dict[str, dict[str, str]] = {}
    manual_keys: set[str] = set()
    for r in (shifts_res.data or []):
        if not r.get("shift"):
            continue
        key = f"{r['nurse_uid']}_{r['date']}"
        if key not in generated_keys:
            manual_keys.add(key)  # 人工預填（非系統生成）→ 淡黃底
        shift_map.setdefault(r["nurse_uid"], {})[r["date"]] = r["shift"]
    nurse_rows = [
        {"uid": u["uid"], "name": u["name"], "level": u.get("level"), "shifts": shift_map.get(u["uid"], {})}
        for u in (users_res.data or [])
    ]

    demand, special_cols = _compute_demand(rules, cycle_dates)
    buf = _build_matrix_excel("完整班表", cycle_dates, nurse_rows, manual_keys,
                              rest_codes=rest_codes, demand=demand, special_date_cols=special_cols)
    filename = f"完整班表_{start_str}_{end_str}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f"attachment; filename*=UTF-8''{quote(filename, safe='')}",
            "Cache-Control": "no-store, no-cache, must-revalidate",
        },
    )


class ExportTempBody(BaseModel):
    schedules: dict[str, dict[str, str]]   # {nurse_uid: {date: shift}}
    cycle_dates: list[str]


@app.post("/export/temp")
def export_temp(
    body: ExportTempBody,
    current_user: dict = Depends(require_roles("admin", "superadmin", "dual")),
):
    """匯出暫時班表（CP-SAT 計算完成但尚未寫入 DB 的結果）"""
    rules_res = supabase.table("rules").select("*").limit(1).execute()
    rules = rules_res.data[0].get("data") or {} if rules_res.data else {}
    cycle = rules.get("cycle", {})
    start_str = body.cycle_dates[0] if body.cycle_dates else cycle.get("start_date")
    end_str   = body.cycle_dates[-1] if body.cycle_dates else cycle.get("end_date")
    if not start_str or not end_str:
        raise HTTPException(400, "無排班日期")

    shift_defs = rules.get("shifts", {})
    rest_codes = {s["code"] for s in shift_defs.get("rest", []) if s.get("code")} or {"OFF", "半"}

    users_res = supabase.table("users").select("uid, name, halftime, level").in_("role", ["nurse", "dual"]).order("sort_order").execute()
    uid_name   = {u["uid"]: u["name"] for u in (users_res.data or [])}
    uid_halftime = {u["uid"]: u.get("halftime", False) for u in (users_res.data or [])}

    # 現有 DB 中的班別（人工預填）→ 淡黃底
    existing_res = supabase.table("shifts").select("nurse_uid, date, shift, confirmed") \
        .gte("date", start_str).lte("date", end_str).execute()
    manual_keys = {
        f"{r['nurse_uid']}_{r['date']}"
        for r in (existing_res.data or []) if r.get("shift")
    }

    shift_map: dict[str, dict[str, str]] = {}
    for uid, date_shifts in body.schedules.items():
        for d_str in body.cycle_dates:
            shift = date_shifts.get(d_str, "")
            if not shift:
                continue
            # 半職護理師的應休班顯示為「休假」
            display = "休假" if (uid_halftime.get(uid) and shift in rest_codes and shift != "OFF") else shift
            shift_map.setdefault(uid, {})[d_str] = display

    nurse_rows = [
        {"uid": u["uid"], "name": u["name"], "level": u.get("level"), "shifts": shift_map.get(u["uid"], {})}
        for u in (users_res.data or [])
    ]

    demand, special_cols = _compute_demand(rules, body.cycle_dates)
    buf = _build_matrix_excel("暫時班表", body.cycle_dates, nurse_rows, manual_keys,
                              rest_codes=rest_codes | {"休假"}, demand=demand, special_date_cols=special_cols)
    filename = f"暫時班表_{start_str}_{end_str}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f"attachment; filename*=UTF-8''{quote(filename, safe='')}",
            "Cache-Control": "no-store, no-cache, must-revalidate",
        },
    )


@app.post("/schedule/clear-generated")
def clear_generated_schedule(
    current_user: dict = Depends(require_roles("admin", "superadmin", "dual")),
):
    """操作1：清除所有 CP-SAT 生成內容，只留下人員填寫的內容"""
    rules_res = supabase.table("rules").select("id", "data").limit(1).execute()
    if not rules_res.data:
        raise HTTPException(400, "找不到規則資料")
    cur_data = rules_res.data[0].get("data") or {}
    rules_id = rules_res.data[0]["id"]

    keys: list[str] = cur_data.get("last_generated_keys", [])
    rng = cur_data.get("last_generated_range")
    if not keys or not rng:
        return {"message": "✓ 無需清除（沒有 CP-SAT 生成紀錄）", "deleted": 0}

    _backup_cycle_shifts(cur_data)  # 執行前備份，供「恢復確認送出及待確認內容」使用

    key_set = set(keys)
    rows_res = supabase.table("shifts").select("id, nurse_uid, date") \
        .gte("date", rng["start_date"]).lte("date", rng["end_date"]).execute()
    ids_to_delete = [
        r["id"] for r in (rows_res.data or [])
        if f"{r['nurse_uid']}_{r['date']}" in key_set
    ]
    if ids_to_delete:
        supabase.table("shifts").delete().in_("id", ids_to_delete).execute()

    cur_data["last_generated_keys"] = []
    supabase.table("rules").update({"data": cur_data}).eq("id", rules_id).execute()

    return {"message": f"✓ 已清除 CP-SAT 生成內容（{len(ids_to_delete)} 格）", "deleted": len(ids_to_delete)}


def _backup_cycle_shifts(cur_data: dict) -> None:
    """把目前預班週期內所有已填寫內容（確認送出＋待確認）備份到 rules.data.manual_backup。
    呼叫端負責把 cur_data 寫回 rules 資料表。"""
    cycle = cur_data.get("cycle", {})
    start_str, end_str = cycle.get("start_date"), cycle.get("end_date")
    if not start_str or not end_str:
        return
    rows_res = supabase.table("shifts").select("nurse_uid, date, shift, confirmed") \
        .gte("date", start_str).lte("date", end_str).execute()
    cur_data["manual_backup"] = {
        "rows": [
            {"nurse_uid": r["nurse_uid"], "date": r["date"],
             "shift": r["shift"], "confirmed": r.get("confirmed", False)}
            for r in (rows_res.data or []) if r.get("shift")
        ],
        "start_date": start_str,
        "end_date": end_str,
        "backed_up_at": datetime.utcnow().isoformat(),
    }


@app.post("/schedule/clear-cycle")
def clear_cycle_schedule(
    current_user: dict = Depends(require_roles("admin", "superadmin", "dual")),
):
    """操作：清除預班週期內所有填寫內容（執行前自動備份，可用「恢復確認送出及待確認內容」還原）"""
    rules_res = supabase.table("rules").select("id", "data").limit(1).execute()
    if not rules_res.data:
        raise HTTPException(400, "找不到規則資料")
    cur_data = rules_res.data[0].get("data") or {}
    rules_id = rules_res.data[0]["id"]

    cycle = cur_data.get("cycle", {})
    start_str, end_str = cycle.get("start_date"), cycle.get("end_date")
    if not start_str or not end_str:
        raise HTTPException(400, "請先設定排班週期")

    _backup_cycle_shifts(cur_data)  # 執行前備份

    rows_res = supabase.table("shifts").select("id") \
        .gte("date", start_str).lte("date", end_str).execute()
    ids = [r["id"] for r in (rows_res.data or [])]
    if ids:
        supabase.table("shifts").delete().in_("id", ids).execute()

    cur_data["last_generated_keys"] = []
    supabase.table("rules").update({"data": cur_data}).eq("id", rules_id).execute()

    return {"message": f"✓ 已清除預班週期內所有填寫內容（{len(ids)} 格），可用「恢復確認送出及待確認內容」還原", "deleted": len(ids)}


@app.post("/schedule/restore-manual")
def restore_manual_schedule(
    current_user: dict = Depends(require_roles("admin", "superadmin", "dual")),
):
    """操作：恢復預班週期內確認送出及待確認的內容（還原最近一次自動備份）"""
    operator_uid = current_user.get("sub")

    rules_res = supabase.table("rules").select("id", "data").limit(1).execute()
    if not rules_res.data:
        raise HTTPException(400, "找不到規則資料")
    cur_data = rules_res.data[0].get("data") or {}

    backup = cur_data.get("manual_backup")
    if not backup or not backup.get("rows"):
        raise HTTPException(400, "找不到備份資料（執行清除/還原操作時才會自動建立備份）")

    start_str, end_str = backup["start_date"], backup["end_date"]

    # 刪除該週期所有現有班別，還原備份（保留原 confirmed 狀態）
    rows_res = supabase.table("shifts").select("id") \
        .gte("date", start_str).lte("date", end_str).execute()
    ids = [r["id"] for r in (rows_res.data or [])]
    if ids:
        supabase.table("shifts").delete().in_("id", ids).execute()

    restore_rows = [
        {
            "code": f"{r['nurse_uid']}_{r['date']}", "label": r["shift"],
            "nurse_uid": r["nurse_uid"], "date": r["date"], "shift": r["shift"],
            "confirmed": r.get("confirmed", False), "updated_by": operator_uid,
        }
        for r in backup["rows"]
    ]
    if restore_rows:
        supabase.table("shifts").insert(restore_rows).execute()

    backed_at = (backup.get("backed_up_at") or "")[:19].replace("T", " ")
    return {
        "message": f"✓ 已恢復確認送出及待確認的內容（{len(restore_rows)} 格，備份時間 {backed_at} UTC）",
        "restored": len(restore_rows),
    }


@app.post("/schedule/restore-generated")
def restore_generated_schedule(
    current_user: dict = Depends(require_roles("admin", "superadmin", "dual")),
):
    """操作2：恢復到上次 CP-SAT 生成的內容"""
    operator_uid = current_user.get("sub")

    rules_res = supabase.table("rules").select("id", "data").limit(1).execute()
    if not rules_res.data:
        raise HTTPException(400, "找不到規則資料")
    cur_data = rules_res.data[0].get("data") or {}
    rules_id = rules_res.data[0]["id"]

    full: dict = cur_data.get("last_generated_full")
    rng = cur_data.get("last_generated_range")
    if not full or not rng:
        raise HTTPException(400, "找不到上次 CP-SAT 生成的完整記錄")

    start_str, end_str = rng["start_date"], rng["end_date"]

    # 執行前備份目前內容，供「恢復確認送出及待確認內容」使用
    _backup_cycle_shifts(cur_data)
    supabase.table("rules").update({"data": cur_data}).eq("id", rules_id).execute()

    # 刪除該週期所有現有班別，還原為上次生成的完整結果
    supabase.table("shifts").delete().gte("date", start_str).lte("date", end_str).execute()

    restore_rows = [
        {
            "code": f"{uid}_{d_str}", "label": shift,
            "nurse_uid": uid, "date": d_str, "shift": shift,
            "confirmed": False, "updated_by": operator_uid,
        }
        for uid, date_shifts in full.items()
        for d_str, shift in date_shifts.items()
        if shift
    ]
    if restore_rows:
        supabase.table("shifts").insert(restore_rows).execute()

    return {
        "message": f"✓ 已恢復到上次 CP-SAT 生成的內容（{len(restore_rows)} 格）",
        "restored": len(restore_rows),
    }


@app.post("/schedule/purge-old")
def purge_old_schedule(
    current_user: dict = Depends(require_roles("admin", "superadmin", "dual")),
):
    """操作3：清除半年之外所有班表（不可復原）"""
    cutoff = (date_type.today() - timedelta(days=182)).isoformat()
    res = supabase.table("shifts").select("id").lt("date", cutoff).execute()
    ids = [r["id"] for r in (res.data or [])]
    if ids:
        supabase.table("shifts").delete().in_("id", ids).execute()
    return {"message": f"✓ 已清除 {cutoff} 之前的班表（{len(ids)} 格）", "deleted": len(ids), "cutoff": cutoff}


@app.get("/logs")
def get_logs(
    limit: int = 200,
    current_user: dict = Depends(require_roles("admin", "superadmin", "dual")),
):
    res = supabase.table("shift_logs").select("*").order("created_at", desc=True).limit(limit).execute()
    return {"logs": res.data}


@app.delete("/logs")
def delete_logs(
    before_hours: Optional[int] = None,
    current_user: dict = Depends(require_roles("admin", "superadmin", "dual")),
):
    if before_hours is None:
        supabase.table("shift_logs").delete().neq("id", "00000000-0000-0000-0000-000000000000").execute()
        return {"message": "已清除所有操作紀錄"}
    cutoff = (datetime.utcnow() - timedelta(hours=before_hours)).isoformat()
    supabase.table("shift_logs").delete().lt("created_at", cutoff).execute()
    return {"message": f"已清除 {before_hours} 小時前的操作紀錄"}
